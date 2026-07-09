"""
AI答案采集平台 - 主应用
"""
import os
import sys

# 抑制 Playwright 的 Node.js 警告
os.environ['NODE_NO_WARNINGS'] = '1'
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from flask import Flask, render_template, request, jsonify, redirect, url_for, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_cors import CORS
from datetime import datetime, timezone, timedelta
import json
import threading
import logging
import time
import platform
from urllib.parse import urlparse
import requests

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(name)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger('app')

# 获取当前北京时间 (CST, UTC+8)
def now_cst():
    return datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=8))).replace(tzinfo=None)

from models import db, User, MonitorTask, CollectionResult, GeoManuscript, SentimentConfig
from config_web import Config
from profile_utils import get_profile_dir, clear_profile_dir
from local_paths import app_data_dir, answers_dir
from version import app_info

try:
    from cloud_sync import (
        cloud_sync_enabled,
        pull_remote_tasks,
        restore_workspace_from_cloud,
        save_cloud_account,
        sync_status,
        sync_user_workspace,
        upload_workspace_assets,
    )
    CLOUD_SYNC_AVAILABLE = True
except Exception as e:
    logger.warning(f"云端同步模块不可用: {e}")
    CLOUD_SYNC_AVAILABLE = False


def _desktop_downloads_dir():
    path = os.path.join(os.path.expanduser('~'), 'Downloads', 'GEO-SOP')
    os.makedirs(path, exist_ok=True)
    return path


def _safe_download_filename(filename):
    cleaned = ''.join(c if c not in '<>:"/\\|?*\n\r\t' else '_' for c in filename)
    return cleaned.strip().strip('.') or 'download'


def _maybe_save_desktop_download(buffer, filename):
    if request.args.get('save_to_downloads') != '1':
        return None
    if os.environ.get('GEO_DESKTOP_MODE') != '1':
        return None

    safe_name = _safe_download_filename(filename)
    target = os.path.join(_desktop_downloads_dir(), safe_name)
    base, ext = os.path.splitext(target)
    counter = 2
    while os.path.exists(target):
        target = f"{base}_{counter}{ext}"
        counter += 1

    current_pos = buffer.tell()
    buffer.seek(0)
    with open(target, 'wb') as f:
        f.write(buffer.read())
    buffer.seek(current_pos)
    return target


def _normalize_openai_chat_url(api_url):
    url = (api_url or '').strip()
    if not url:
        return url
    normalized = url.rstrip('/')
    if normalized.endswith('/chat/completions'):
        return normalized
    if normalized.endswith('/v1'):
        return f'{normalized}/chat/completions'
    parsed = urlparse(normalized)
    if parsed.netloc == 'api.openai.com' and parsed.path in ('', '/'):
        return f'{normalized}/v1/chat/completions'
    return f'{normalized}/chat/completions'


def _normalize_anthropic_messages_url(api_url):
    url = (api_url or '').strip()
    if not url:
        return url
    normalized = url.rstrip('/')
    if normalized.endswith('/v1/messages'):
        return normalized
    if normalized.endswith('/v1'):
        return f'{normalized}/messages'
    return f'{normalized}/v1/messages'


def _ai_api_mode(config):
    mode = (config.ai_platform or 'openai').strip().lower()
    return 'anthropic' if mode == 'anthropic' else 'openai'


def _version_parts(version):
    base = str(version or '').lstrip('v').split('-', 1)[0]
    parts = []
    for item in base.split('.'):
        try:
            parts.append(int(item))
        except ValueError:
            parts.append(0)
    while len(parts) < 3:
        parts.append(0)
    return parts[:3]


def _check_latest_update():
    info = app_info()
    update_url = app.config.get('GEO_UPDATE_URL') or os.environ.get('GEO_UPDATE_URL') or 'https://geo.allgood.cn/update.json'
    current_version = info.get('version') or ''
    status = {
        'current_version': current_version,
        'latest_version': current_version,
        'has_update': False,
        'update_url': update_url,
    }
    try:
        response = requests.get(update_url, timeout=8)
        response.raise_for_status()
        manifest = response.json()
        system_name = platform.system().lower()
        platform_key = 'windows' if system_name.startswith('win') else 'macos' if system_name == 'darwin' else 'linux'
        downloads = manifest.get('downloads') if isinstance(manifest.get('downloads'), dict) else {}
        package = downloads.get(platform_key) if isinstance(downloads.get(platform_key), dict) else {}
        latest_version = str(package.get('version') or manifest.get('version') or current_version)
        status.update({
            'latest_version': latest_version,
            'has_update': _version_parts(latest_version) > _version_parts(current_version),
            'channel': manifest.get('channel'),
            'released_at': manifest.get('released_at'),
            'notes': manifest.get('notes') or [],
            'download_url': package.get('url'),
            'download_name': package.get('name'),
            'download_size': package.get('size'),
            'platform': platform_key,
        })
    except Exception as e:
        status['error'] = str(e)
    return status


def _extract_json_text(text):
    cleaned = (text or '').strip()
    if cleaned.startswith('```json'):
        cleaned = cleaned[7:]
    elif cleaned.startswith('```'):
        cleaned = cleaned[3:]
    if cleaned.endswith('```'):
        cleaned = cleaned[:-3]
    return cleaned.strip()


def _resolve_screenshot_path(filepath):
    if not filepath:
        return None

    raw_path = filepath.replace('\\', '/')
    absolute_path = os.path.normpath(raw_path) if os.path.isabs(raw_path) else None
    normalized = raw_path.lstrip('/')
    without_answers = normalized[len('answers/'):] if normalized.startswith('answers/') else normalized

    webapp_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir = os.path.dirname(webapp_dir)
    local_answers_dir = answers_dir()

    candidates = [
        absolute_path,
        os.path.normpath(normalized),
        os.path.normpath(os.path.join(local_answers_dir, normalized)),
        os.path.normpath(os.path.join(local_answers_dir, without_answers)),
        os.path.normpath(os.path.join(webapp_dir, normalized)),
        os.path.normpath(os.path.join(root_dir, normalized)),
        os.path.normpath(os.path.join(webapp_dir, 'answers', without_answers)),
        os.path.normpath(os.path.join(root_dir, 'answers', without_answers)),
    ]
    allowed_roots = [
        os.path.normpath(local_answers_dir),
        os.path.normpath(os.path.join(webapp_dir, 'answers')),
        os.path.normpath(os.path.join(root_dir, 'answers')),
    ]

    for candidate in filter(None, candidates):
        if any(candidate.startswith(root) for root in allowed_roots) and os.path.isfile(candidate):
            return candidate
    return None


def _login_status_cache_path():
    return os.path.join(app_data_dir(), 'login_status_cache.json')


def _read_login_status_cache():
    path = _login_status_cache_path()
    if not os.path.isfile(path):
        return {}
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception as e:
        logger.warning(f"[LoginCache] 读取登录状态缓存失败: {e}")
        return {}


def _write_login_status_cache(data):
    path = _login_status_cache_path()
    tmp_path = f"{path}.tmp"
    try:
        with open(tmp_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, path)
    except Exception as e:
        logger.warning(f"[LoginCache] 写入登录状态缓存失败: {e}")


_cloud_sync_lock = threading.Lock()
_cloud_sync_running_users = set()


def _sync_user_workspace_blocking(user_id, reason='workspace_change', wait=False):
    """同步当前用户工作区；后台采集完成时用 wait=True 避免跳过最终结果。"""
    if not CLOUD_SYNC_AVAILABLE:
        return False
    if not cloud_sync_enabled():
        return False

    user_id = int(user_id)
    while True:
        with _cloud_sync_lock:
            if user_id not in _cloud_sync_running_users:
                _cloud_sync_running_users.add(user_id)
                break
        if not wait:
            return False
        time.sleep(1)

    try:
        with app.app_context():
            result = sync_user_workspace(user_id)
            logger.info("[CloudSync] %s user=%s result=%s", reason, user_id, result)
        return True
    except Exception as e:
        logger.exception("[CloudSync] 同步失败 user=%s reason=%s: %s", user_id, reason, e)
        return False
    finally:
        with _cloud_sync_lock:
            _cloud_sync_running_users.discard(user_id)


def _queue_cloud_sync(user_id, reason='workspace_change'):
    """后台同步当前用户工作区到云端 MySQL 镜像库。"""
    if not CLOUD_SYNC_AVAILABLE:
        return False
    if not cloud_sync_enabled():
        return False

    def worker():
        _sync_user_workspace_blocking(user_id, reason, wait=False)

    threading.Thread(target=worker, daemon=True).start()
    return True


def _adopt_local_workspace_for_cloud_user(user):
    """首次云端登录时，把旧单机 local 工作区迁移到当前云端账号。"""
    local_user = User.query.filter_by(username='local').first()
    if not local_user or local_user.id == user.id:
        return 0

    changed = 0
    for task in MonitorTask.query.filter_by(user_id=local_user.id).all():
        task.user_id = user.id
        changed += 1
    for manuscript in GeoManuscript.query.filter_by(user_id=local_user.id).all():
        manuscript.user_id = user.id
        changed += 1
    for config in SentimentConfig.query.filter_by(user_id=local_user.id).all():
        config.user_id = user.id
        changed += 1

    if changed:
        db.session.commit()
        logger.info("[CloudSync] adopted %s local workspace records for cloud user=%s", changed, user.id)
    return changed


def _restore_cloud_workspace_if_empty(user):
    """新安装桌面端登录云端账号后，自动恢复云端历史记录。"""
    if not CLOUD_SYNC_AVAILABLE or not cloud_sync_enabled():
        return {'restored': False, 'enabled': False}
    try:
        result = restore_workspace_from_cloud(user.id, only_if_empty=True)
        logger.info("[CloudSync] restore user=%s result=%s", user.id, result)
        return result
    except Exception as e:
        logger.exception("[CloudSync] 自动恢复云端历史失败 user=%s: %s", user.id, e)
        return {'restored': False, 'error': str(e)}


def _get_cached_login_status(user_id):
    cache = _read_login_status_cache()
    user_cache = cache.get(str(user_id), {})
    return user_cache if isinstance(user_cache, dict) else {}


def _normalize_cached_login_status(value):
    if isinstance(value, dict):
        return {
            'is_logged_in': bool(value.get('is_logged_in', value.get('logged_in', False))),
            'checked_at': value.get('checked_at'),
            'error': value.get('error')
        }
    return {
        'is_logged_in': bool(value),
        'checked_at': None,
        'error': None
    }


def _set_cached_login_status(user_id, platform, is_logged_in, error=None):
    cache = _read_login_status_cache()
    user_key = str(user_id)
    if not isinstance(cache.get(user_key), dict):
        cache[user_key] = {}
    cache[user_key][platform] = {
        'is_logged_in': bool(is_logged_in),
        'checked_at': now_cst().strftime('%Y-%m-%d %H:%M:%S'),
        'error': error
    }
    _write_login_status_cache(cache)

# 导入登录检测模块
try:
    from login_checker import check_platform_login, check_all_platforms
    LOGIN_CHECKER_AVAILABLE = True
except ImportError:
    LOGIN_CHECKER_AVAILABLE = False
    print("警告: login_checker 模块不可用，登录检测功能将被禁用")

# 导入登录辅助模块
try:
    from login_helper import open_login_browser
    LOGIN_HELPER_AVAILABLE = True
    print("OK: login_helper 模块导入成功")
except ImportError as e:
    LOGIN_HELPER_AVAILABLE = False
    print(f"警告: login_helper 模块不可用，浏览器登录功能将被禁用")
    print(f"      错误详情: {e}")
except Exception as e:
    LOGIN_HELPER_AVAILABLE = False
    print(f"警告: login_helper 模块加载失败，浏览器登录功能将被禁用")
    print(f"      错误详情: {type(e).__name__}: {e}")

# 导入调度器模块
try:
    from scheduler import init_scheduler, add_task_job, remove_task_job
    SCHEDULER_AVAILABLE = True
except ImportError:
    SCHEDULER_AVAILABLE = False
    print("警告: scheduler 模块不可用，定时任务功能将被禁用")

app = Flask(__name__)
app.config.from_object(Config)

# 初始化扩展
db.init_app(app)
CORS(app)

# 初始化定时调度器
if SCHEDULER_AVAILABLE and not Config.DESKTOP_MODE:
    # 检查是否在reloader进程中，避免调度器被多次启动
    import os
    if os.environ.get('WERKZEUG_RUN_MAIN'):
        # 主进程：初始化调度器
        init_scheduler(app)
    elif not os.environ.get('FLASK_APP'):
        # 非Flask reloader模式（如直接运行）：也初始化调度器
        init_scheduler(app)
    else:
        print("  检测到Werkzeug reloader进程，跳过调度器初始化（将在主进程中启动）")

# 登录管理
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


@app.before_request
def desktop_auto_login():
    if not app.config.get('DESKTOP_MODE'):
        return
    if app.config.get('REQUIRE_LOGIN') and current_user.is_authenticated:
        if current_user.username == 'local' or current_user.email == 'local@geo-sop.local':
            logout_user()
            if request.endpoint not in {'login', 'static'}:
                return redirect(url_for('login'))
        return
    if app.config.get('REQUIRE_LOGIN'):
        return
    if request.endpoint == 'static':
        return
    if current_user.is_authenticated:
        return

    user = User.query.filter_by(username='local').first()
    if not user:
        user = User(username='local', email='local@geo-sop.local')
        user.set_password(os.environ.get('GEO_DESKTOP_LOCAL_PASSWORD', 'local-only'))
        db.session.add(user)
        db.session.commit()
    login_user(user)


@app.after_request
def auto_cloud_sync_after_mutation(response):
    if (
        CLOUD_SYNC_AVAILABLE
        and cloud_sync_enabled()
        and response.status_code < 400
        and request.method in {'POST', 'PUT', 'PATCH', 'DELETE'}
        and request.path.startswith('/api/')
        and not request.path.startswith('/api/cloud-sync')
        and current_user.is_authenticated
    ):
        _queue_cloud_sync(current_user.id, f'{request.method} {request.path}')
    return response


# ==================== 路由 ====================

@app.route('/')
def index():
    """首页 - 重定向到登录或仪表板"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    """用户登录"""
    if app.config.get('DESKTOP_MODE') and not app.config.get('REQUIRE_LOGIN') and request.method == 'GET':
        return redirect(url_for('dashboard', welcome=1))

    if request.method == 'POST':
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        cloud_login = bool(data.get('cloud_login'))

        if cloud_login:
            try:
                cloud_url = (app.config.get('CLOUD_SYNC_URL') or 'https://geo.allgood.cn/api').rstrip('/')
                response = requests.post(
                    f'{cloud_url}/auth/login/',
                    json={'account': username, 'password': password},
                    timeout=20
                )
                payload = response.json()
                if response.status_code >= 400 or not payload.get('success'):
                    return jsonify({'success': False, 'message': payload.get('message', '云端账号登录失败')}), 401

                cloud_user = payload.get('user') or {}
                cloud_username = cloud_user.get('username') or username
                cloud_email = cloud_user.get('email') or f'{cloud_username}@geo.allgood.cn'
                user = User.query.filter((User.username == cloud_username) | (User.email == cloud_email)).first()
                if not user:
                    user = User(username=cloud_username, email=cloud_email)
                    db.session.add(user)
                user.set_password(password)
                db.session.commit()
                adopted_count = _adopt_local_workspace_for_cloud_user(user)

                cloud_sync_url_value = payload.get('cloud_sync_url') or cloud_url
                cloud_sync_token_value = payload.get('token') or ''
                save_cloud_account({
                    'cloud_sync_url': cloud_sync_url_value,
                    'token': cloud_sync_token_value,
                    'user': cloud_user,
                    'logged_in_at': now_cst().strftime('%Y-%m-%d %H:%M:%S')
                })
                os.environ['GEO_CLOUD_SYNC_ENABLED'] = '1'
                os.environ['GEO_CLOUD_SYNC_URL'] = cloud_sync_url_value
                os.environ['GEO_CLOUD_SYNC_TOKEN'] = cloud_sync_token_value
                app.config['CLOUD_SYNC_ENABLED'] = True
                app.config['CLOUD_SYNC_URL'] = cloud_sync_url_value
                app.config['CLOUD_SYNC_TOKEN'] = cloud_sync_token_value

                login_user(user)
                restore_result = {'restored': False}
                if not adopted_count:
                    restore_result = _restore_cloud_workspace_if_empty(user)
                _queue_cloud_sync(user.id, 'cloud_login_adopted' if adopted_count else 'cloud_login')
                message = '云端账号登录成功'
                if restore_result.get('restored'):
                    counts = restore_result.get('counts') or {}
                    message = f"云端账号登录成功，已恢复历史记录：{counts.get('tasks', 0)} 个任务、{counts.get('results', 0)} 条结果"
                return jsonify({'success': True, 'message': message, 'restore': restore_result})
            except Exception as e:
                logger.exception("云端账号登录失败: %s", e)
                return jsonify({'success': False, 'message': f'云端账号登录失败: {str(e)}'}), 500
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user)
            return jsonify({'success': True, 'message': '登录成功'})
        
        return jsonify({'success': False, 'message': '用户名或密码错误'}), 401
    
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    """用户注册"""
    if request.method == 'POST':
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        email = data.get('email')
        
        # 检查用户是否已存在
        if User.query.filter_by(username=username).first():
            return jsonify({'success': False, 'message': '用户名已存在'}), 400
        
        if User.query.filter_by(email=email).first():
            return jsonify({'success': False, 'message': '邮箱已被注册'}), 400
        
        # 创建新用户
        user = User(username=username, email=email)
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        login_user(user)
        _queue_cloud_sync(user.id, 'register')
        
        return jsonify({'success': True, 'message': '注册成功', 'redirect': url_for('dashboard', welcome=1)})
    
    return render_template('register.html')


@app.route('/logout')
@login_required
def logout():
    """用户登出"""
    logout_user()
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    """仪表板 - 监控任务列表"""
    return render_template('dashboard.html')


@app.route('/api/current-user', methods=['GET'])
@login_required
def get_current_user():
    """获取当前登录用户，用于前端按用户隔离本地缓存和首次引导"""
    if app.config.get('DESKTOP_MODE') and current_user.username != 'local':
        _adopt_local_workspace_for_cloud_user(current_user)
    try:
        cloud_status = sync_status(current_user.id) if CLOUD_SYNC_AVAILABLE else {'enabled': False}
    except Exception as e:
        logger.warning("[CloudSync] 当前用户同步状态读取失败: %s", e)
        cloud_status = {
            'enabled': app.config.get('CLOUD_SYNC_ENABLED', False),
            'api_configured': bool(app.config.get('CLOUD_SYNC_URL')),
            'token_configured': bool(app.config.get('CLOUD_SYNC_TOKEN')),
            'error': str(e)
        }
    return jsonify({
        'success': True,
        'user': current_user.to_dict(),
        'desktop_mode': app.config.get('DESKTOP_MODE', False),
        'require_login': app.config.get('REQUIRE_LOGIN', False),
        'cloud_sync': cloud_status,
        'app': app_info(),
        'update': _check_latest_update() if app.config.get('DESKTOP_MODE', False) else {'has_update': False},
        'data_dir': app_data_dir() if app.config.get('DESKTOP_MODE', False) else app.config.get('DATA_DIR')
    })


@app.route('/api/app-info', methods=['GET'])
@login_required
def get_app_info():
    """获取应用版本、运行模式和本地数据目录。"""
    desktop_mode = app.config.get('DESKTOP_MODE', False)
    return jsonify({
        'success': True,
        'app': app_info(),
        'desktop_mode': desktop_mode,
        'data_dir': app_data_dir() if desktop_mode else app.config.get('DATA_DIR'),
        'answers_dir': app.config.get('ANSWERS_DIR'),
        'update': _check_latest_update() if desktop_mode else {'has_update': False}
    })


@app.route('/api/cloud-sync/status', methods=['GET'])
@login_required
def get_cloud_sync_status():
    """查看当前用户云端镜像同步状态。"""
    if not CLOUD_SYNC_AVAILABLE:
        return jsonify({'success': True, 'cloud_sync': {'enabled': False, 'message': '同步模块不可用'}})
    try:
        return jsonify({'success': True, 'cloud_sync': sync_status(current_user.id)})
    except Exception as e:
        logger.exception("[CloudSync] 状态检查失败: %s", e)
        return jsonify({
            'success': False,
            'message': '云端同步状态检查失败',
            'error': str(e)
        }), 500


@app.route('/api/cloud-sync/run', methods=['POST'])
@login_required
def run_cloud_sync_now():
    """立即同步当前用户工作区到云端 MySQL 镜像库。"""
    if not CLOUD_SYNC_AVAILABLE or not cloud_sync_enabled():
        return jsonify({'success': False, 'message': '云端同步未启用'}), 400
    try:
        result = sync_user_workspace(current_user.id)
        return jsonify({'success': True, 'cloud_sync': result})
    except Exception as e:
        logger.exception("[CloudSync] 手动同步失败: %s", e)
        return jsonify({
            'success': False,
            'message': '云端同步失败',
            'error': str(e)
        }), 500


@app.route('/api/cloud-sync/upload-assets', methods=['POST'])
@login_required
def upload_cloud_assets_now():
    """一键上传统计数据和本地截图文件到云端。"""
    if not CLOUD_SYNC_AVAILABLE or not cloud_sync_enabled():
        return jsonify({'success': False, 'message': '云端同步未启用'}), 400
    try:
        sync_result = sync_user_workspace(current_user.id)
        assets_result = upload_workspace_assets(current_user.id, _resolve_screenshot_path)
        return jsonify({
            'success': True,
            'message': '云端上传完成',
            'cloud_sync': sync_result,
            'assets': assets_result
        })
    except Exception as e:
        logger.exception("[CloudSync] 上传云端资源失败: %s", e)
        return jsonify({
            'success': False,
            'message': '上传云端失败',
            'error': str(e)
        }), 500


@app.route('/api/cloud-sync/pull-tasks', methods=['POST'])
@login_required
def pull_cloud_remote_tasks():
    """从服务器端拉取该用户创建的远程采集任务，写入本地任务表。"""
    if not CLOUD_SYNC_AVAILABLE or not cloud_sync_enabled():
        return jsonify({'success': False, 'message': '云端同步未启用'}), 400
    try:
        result = pull_remote_tasks(current_user.id)
        if result.get('created'):
            _queue_cloud_sync(current_user.id, 'pull_remote_tasks')
        return jsonify({'success': True, 'remote_tasks': result})
    except Exception as e:
        logger.exception("[CloudSync] 拉取远程任务失败: %s", e)
        return jsonify({
            'success': False,
            'message': '拉取远程任务失败',
            'error': str(e)
        }), 500


@app.route('/api/platforms', methods=['GET'])
def get_platforms():
    """获取所有支持的AI平台列表"""
    platforms = [
        {'id': 'doubao', 'name': '豆包', 'url': 'https://www.doubao.com/chat'},
        {'id': 'deepseek', 'name': 'DeepSeek', 'url': 'https://chat.deepseek.com'},
        {'id': 'yuanbao', 'name': '元宝', 'url': 'https://yuanbao.tencent.com/chat'},
        {'id': 'kimi', 'name': 'Kimi', 'url': 'https://www.kimi.com'},
        {'id': 'qianwen', 'name': '千问', 'url': 'https://www.qianwen.com'},
        {'id': 'wenxin', 'name': '文心一言(wenxin)', 'url': 'https://wenxin.baidu.com'},
        {'id': 'yiyan', 'name': '文心一言(yiyan)', 'url': 'https://yiyan.baidu.com'},
        {'id': 'baidu_chat', 'name': '文心（chat）', 'url': 'https://chat.baidu.com'},
        {'id': 'chatgpt', 'name': 'ChatGPT', 'url': 'https://chatgpt.com'}
    ]
    return jsonify({
        'success': True,
        'platforms': platforms
    })


@app.route('/api/tasks', methods=['GET'])
@login_required
def get_tasks():
    """获取当前用户的所有监控任务"""
    tasks = MonitorTask.query.filter_by(user_id=current_user.id).order_by(MonitorTask.created_at.desc()).all()
    
    return jsonify({
        'success': True,
        'tasks': [task.to_dict() for task in tasks]
    })


@app.route('/api/tasks', methods=['POST'])
@login_required
def create_task():
    """创建新的监控任务"""
    data = request.get_json()
    
    # 验证必填字段
    required_fields = ['name', 'brand_keywords', 'questions', 'platforms']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'success': False, 'message': f'缺少必填字段: {field}'}), 400
    
    # 创建任务
    schedule_enabled = data.get('schedule_enabled', False)
    
    # 验证舆情配置ID（如果提供了）
    sentiment_config_id = data.get('sentiment_config_id')
    if sentiment_config_id:
        config = SentimentConfig.query.filter_by(id=sentiment_config_id, user_id=current_user.id).first()
        if not config:
            return jsonify({'success': False, 'message': '无效的舆情配置ID'}), 400
    
    # 验证并规范化 max_parallel_platforms
    max_parallel = data.get('max_parallel_platforms', 3)
    if max_parallel is None or not isinstance(max_parallel, int) or max_parallel < 1:
        max_parallel = 3  # 默认值
    
    task = MonitorTask(
        user_id=current_user.id,
        name=data['name'],
        brand_name=data.get('brand_name', ''),
        brand_keywords=json.dumps(data['brand_keywords'], ensure_ascii=False),
        competitor_brands=json.dumps(data.get('competitor_brands', []), ensure_ascii=False),
        questions=json.dumps(data['questions'], ensure_ascii=False),
        platforms=json.dumps(data['platforms'], ensure_ascii=False),
        max_parallel_platforms=max_parallel,
        screenshot_config=json.dumps(data.get('screenshot_config', {}), ensure_ascii=False),
        schedule_type=data.get('schedule_type', 'manual'),
        schedule_config=json.dumps(data.get('schedule_config', {}), ensure_ascii=False),
        schedule_enabled=schedule_enabled,
        sentiment_config_id=sentiment_config_id,
        status='pending'
    )
    
    db.session.add(task)
    db.session.commit()
    
    # 如果启用了定时调度，添加到调度器
    if SCHEDULER_AVAILABLE and schedule_enabled and task.schedule_type in ['daily', 'weekly']:
        add_task_job(app, task.id)
    
    return jsonify({
        'success': True,
        'message': '任务创建成功',
        'task': task.to_dict()
    })


@app.route('/api/tasks/<int:task_id>', methods=['GET'])
@login_required
def get_task(task_id):
    """获取单个任务详情"""
    task = db.session.get(MonitorTask, task_id)
    if not task or task.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权访问'}), 403
    
    task_dict = task.to_dict()
    
    # 如果有绑定的舆情配置，获取配置名称
    if task.sentiment_config_id:
        config = SentimentConfig.query.filter_by(id=task.sentiment_config_id, user_id=current_user.id).first()
        if config:
            task_dict['sentiment_config_name'] = config.name
    
    return jsonify({
        'success': True,
        'task': task_dict
    })


@app.route('/api/tasks/<int:task_id>', methods=['PUT'])
@login_required
def update_task(task_id):
    """更新任务"""
    task = db.session.get(MonitorTask, task_id)
    if not task or task.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权访问'}), 403
    
    data = request.get_json()
    
    # 更新字段
    if 'name' in data:
        task.name = data['name']
    if 'brand_name' in data:
        task.brand_name = data['brand_name']
    if 'brand_keywords' in data:
        task.brand_keywords = json.dumps(data['brand_keywords'], ensure_ascii=False)
    if 'competitor_brands' in data:
        task.competitor_brands = json.dumps(data['competitor_brands'], ensure_ascii=False)
    if 'questions' in data:
        task.questions = json.dumps(data['questions'], ensure_ascii=False)
    if 'platforms' in data:
        task.platforms = json.dumps(data['platforms'], ensure_ascii=False)
    if 'max_parallel_platforms' in data:
        # 验证并规范化 max_parallel_platforms
        max_parallel = data['max_parallel_platforms']
        if max_parallel is None or not isinstance(max_parallel, int) or max_parallel < 1:
            max_parallel = 3  # 默认值
        task.max_parallel_platforms = max_parallel
    if 'screenshot_config' in data:
        task.screenshot_config = json.dumps(data['screenshot_config'], ensure_ascii=False)
    if 'schedule_type' in data:
        task.schedule_type = data['schedule_type']
    if 'schedule_config' in data:
        task.schedule_config = json.dumps(data['schedule_config'], ensure_ascii=False)
    if 'schedule_enabled' in data:
        task.schedule_enabled = data['schedule_enabled']
    if 'sentiment_config_id' in data:
        # 验证舆情配置ID
        if data['sentiment_config_id']:
            config = SentimentConfig.query.filter_by(id=data['sentiment_config_id'], user_id=current_user.id).first()
            if not config:
                return jsonify({'success': False, 'message': '无效的舆情配置ID'}), 400
        task.sentiment_config_id = data['sentiment_config_id']
    
    db.session.commit()
    
    # 更新调度器任务
    if SCHEDULER_AVAILABLE and task.schedule_type in ['daily', 'weekly']:
        if task.schedule_enabled:
            add_task_job(app, task.id)
        else:
            remove_task_job(task.id)
    
    return jsonify({
        'success': True,
        'message': '任务更新成功',
        'task': task.to_dict()
    })


@app.route('/api/tasks/<int:task_id>', methods=['DELETE'])
@login_required
def delete_task(task_id):
    """删除任务"""
    task = db.session.get(MonitorTask, task_id)
    if not task or task.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权访问'}), 403
    
    # 从调度器中移除
    if SCHEDULER_AVAILABLE:
        remove_task_job(task_id)
    
    db.session.delete(task)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': '任务删除成功'
    })


@app.route('/api/tasks/<int:task_id>/run', methods=['POST'])
@login_required
def run_task(task_id):
    """执行任务 - 开始采集数据"""
    task = db.session.get(MonitorTask, task_id)
    if not task or task.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权访问'}), 403
    
    # 获取前端传来的全局间隔设置
    data = request.get_json() or {}
    min_interval = data.get('min_interval')
    max_interval = data.get('max_interval')
    
    # 检查任务状态
    if task.status == 'running':
        return jsonify({'success': False, 'message': '任务正在执行中'}), 400
    if task.status == 'paused':
        return jsonify({'success': False, 'message': '任务已暂停，请点击继续采集'}), 400
    
    # 更新任务状态为执行中
    task_user_id = int(task.user_id)
    task.status = 'running'
    task.last_run_at = now_cst()
    db.session.commit()
    
    # 在后台线程中执行采集
    import threading
    def run_in_background():
        from collector import run_collection
        try:
            run_collection(task_id, min_interval=min_interval, max_interval=max_interval)
        except Exception as e:
            print(f"采集失败: {e}")
            with app.app_context():
                task = db.session.get(MonitorTask, task_id)
                if task:
                    task.status = 'failed'
                    db.session.commit()
        finally:
            _sync_user_workspace_blocking(task_user_id, 'task_run_finished', wait=True)
    
    thread = threading.Thread(target=run_in_background)
    thread.daemon = True
    thread.start()
    
    return jsonify({
        'success': True,
        'message': '任务已开始执行'
    })


@app.route('/api/tasks/<int:task_id>/control', methods=['POST'])
@login_required
def control_task(task_id):
    """控制任务执行：pause / resume / stop"""
    task = db.session.get(MonitorTask, task_id)
    if not task or task.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权访问'}), 403

    data = request.get_json() or {}
    command = (data.get('command') or '').strip().lower()

    if command not in ['pause', 'resume', 'stop']:
        return jsonify({'success': False, 'message': '无效控制命令'}), 400

    if command == 'pause':
        if task.status != 'running':
            return jsonify({'success': False, 'message': '仅执行中的任务可暂停'}), 400
        task.control_command = 'pause'
        db.session.commit()
        return jsonify({
            'success': True,
            'message': '已发送暂停命令，将在当前问题采集完成后暂停',
            'task': task.to_dict()
        })

    if command == 'resume':
        if task.status != 'paused':
            return jsonify({'success': False, 'message': '仅暂停中的任务可继续'}), 400
        task.control_command = 'resume'
        db.session.commit()
        return jsonify({
            'success': True,
            'message': '已发送继续命令，任务将恢复采集',
            'task': task.to_dict()
        })

    if task.status != 'paused':
        return jsonify({'success': False, 'message': '仅暂停中的任务可结束'}), 400

    task.control_command = 'stop'
    db.session.commit()
    return jsonify({
        'success': True,
        'message': '已发送结束命令，任务将停止采集',
        'task': task.to_dict()
    })


@app.route('/api/tasks/<int:task_id>/reset-status', methods=['POST'])
@login_required
def reset_task_status(task_id):
    """重置任务状态（用于中断后恢复）"""
    task = db.session.get(MonitorTask, task_id)
    if not task or task.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权访问'}), 403
    
    # 只有 running 或 failed 状态可以重置
    if task.status not in ['running', 'failed']:
        return jsonify({'success': False, 'message': '只有执行中或失败的任务可以重置'}), 400
    
    task.status = 'pending'
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': '状态已重置，可以重新执行'
    })


@app.route('/api/tasks/<int:task_id>/results', methods=['GET'])
@login_required
def get_task_results(task_id):
    """获取任务的采集结果"""
    task = db.session.get(MonitorTask, task_id)
    if not task or task.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权访问'}), 403
    
    # 获取所有采集结果
    results = CollectionResult.query.filter_by(task_id=task_id).order_by(CollectionResult.created_at.desc()).all()
    
    # 计算品牌词曝光率
    brand_keywords = json.loads(task.brand_keywords)
    exposure_stats = calculate_exposure_stats(results, brand_keywords)
    
    # 获取舆情配置：只使用任务绑定的配置（不自动使用默认配置）
    sentiment_config = None
    if task.sentiment_config_id:
        # 使用任务绑定的舆情配置
        sentiment_config = SentimentConfig.query.filter_by(id=task.sentiment_config_id, user_id=current_user.id).first()
    
    # 对每个结果进行智能舆情分析
    results_with_sentiment = []
    for result in results:
        result_dict = result.to_dict()
        
        # 检查是否已有缓存的智能舆情分析结果
        cached_sentiment = result.ai_sentiment_result
        
        # 只有当任务绑定了启用智能舆情的配置时，才进行分析
        if sentiment_config and sentiment_config.enable_ai_sentiment and sentiment_config.ai_api_url and sentiment_config.ai_api_key:
            if cached_sentiment:
                # 使用缓存的分析结果
                result_dict['ai_sentiment'] = json.loads(cached_sentiment)
                result_dict['ai_sentiment_cached'] = True
                result_dict['ai_sentiment_updated_at'] = result.ai_sentiment_updated_at.strftime('%Y-%m-%d %H:%M:%S') if result.ai_sentiment_updated_at else None
            else:
                # 首次分析，调用AI接口
                ai_sentiment = analyze_sentiment_ai(result.answer or '', '', sentiment_config)
                
                # 如果智能舆情分析失败（包含错误信息），降级到普通舆情分析
                if ai_sentiment.get('error') or '分析失败' in ai_sentiment.get('reason', ''):
                    positive_words = json.loads(sentiment_config.positive_words) if sentiment_config.positive_words else []
                    negative_words = json.loads(sentiment_config.negative_words) if sentiment_config.negative_words else []
                    local_sentiment = analyze_sentiment_local(result.answer or '', '', positive_words, negative_words)
                    ai_sentiment = {
                        **local_sentiment,
                        'reason': f"智能分析失败，已使用关键词分析: {local_sentiment.get('label')}"
                    }
                
                # 保存分析结果到数据库
                result.ai_sentiment_result = json.dumps(ai_sentiment)
                result.ai_sentiment_updated_at = now_cst()
                db.session.commit()
                
                result_dict['ai_sentiment'] = ai_sentiment
                result_dict['ai_sentiment_cached'] = False
            
            result_dict['sentiment_config_name'] = sentiment_config.name  # 添加配置名称用于显示
        elif sentiment_config and (sentiment_config.positive_words or sentiment_config.negative_words):
            # 如果没有启用智能舆情但有关键词配置，使用普通舆情分析
            if cached_sentiment:
                result_dict['ai_sentiment'] = json.loads(cached_sentiment)
                result_dict['ai_sentiment_cached'] = True
                result_dict['ai_sentiment_updated_at'] = result.ai_sentiment_updated_at.strftime('%Y-%m-%d %H:%M:%S') if result.ai_sentiment_updated_at else None
            else:
                positive_words = json.loads(sentiment_config.positive_words) if sentiment_config.positive_words else []
                negative_words = json.loads(sentiment_config.negative_words) if sentiment_config.negative_words else []
                local_sentiment = analyze_sentiment_local(result.answer or '', '', positive_words, negative_words)
                
                # 保存分析结果到数据库
                result.ai_sentiment_result = json.dumps(local_sentiment)
                result.ai_sentiment_updated_at = now_cst()
                db.session.commit()
                
                result_dict['ai_sentiment'] = local_sentiment
                result_dict['ai_sentiment_cached'] = False
            
            result_dict['sentiment_config_name'] = sentiment_config.name
        else:
            result_dict['ai_sentiment'] = None
            result_dict['sentiment_config_name'] = sentiment_config.name if sentiment_config else None
        
        results_with_sentiment.append(result_dict)
    
    task_dict = task.to_dict()
    if task.sentiment_config_id and sentiment_config:
        task_dict['sentiment_config_name'] = sentiment_config.name
    
    return jsonify({
        'success': True,
        'task': task_dict,
        'results': results_with_sentiment,
        'exposure_stats': exposure_stats
    })


@app.route('/api/results/<int:result_id>', methods=['GET'])
@login_required
def get_result(result_id):
    """获取单个采集结果详情"""
    result = db.session.get(CollectionResult, result_id)
    if not result:
        return jsonify({'success': False, 'message': '结果不存在'}), 404
    
    task = db.session.get(MonitorTask, result.task_id)
    if not task or task.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权访问'}), 403
    
    return jsonify({
        'success': True,
        'result': result.to_dict()
    })


@app.route('/api/results/<int:result_id>/update-sentiment', methods=['POST'])
@login_required
def update_sentiment(result_id):
    """更新单条采集结果的智能舆情分析"""
    result = db.session.get(CollectionResult, result_id)
    if not result:
        return jsonify({'success': False, 'message': '结果不存在'}), 404
    
    task = db.session.get(MonitorTask, result.task_id)
    if not task or task.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权访问'}), 403
    
    # 获取舆情配置
    sentiment_config = None
    if task.sentiment_config_id:
        sentiment_config = SentimentConfig.query.filter_by(id=task.sentiment_config_id, user_id=current_user.id).first()
    
    if not sentiment_config:
        return jsonify({'success': False, 'message': '未绑定舆情配置'}), 400
    
    # 重新分析舆情
    ai_sentiment = None
    if sentiment_config.enable_ai_sentiment and sentiment_config.ai_api_url and sentiment_config.ai_api_key:
        ai_sentiment = analyze_sentiment_ai(result.answer or '', '', sentiment_config)
        
        # 如果智能舆情分析失败，降级到普通舆情分析
        if ai_sentiment.get('error') or '分析失败' in ai_sentiment.get('reason', ''):
            positive_words = json.loads(sentiment_config.positive_words) if sentiment_config.positive_words else []
            negative_words = json.loads(sentiment_config.negative_words) if sentiment_config.negative_words else []
            local_sentiment = analyze_sentiment_local(result.answer or '', '', positive_words, negative_words)
            ai_sentiment = {
                **local_sentiment,
                'reason': f"智能分析失败，已使用关键词分析: {local_sentiment.get('label')}"
            }
    else:
        # 使用普通舆情分析
        positive_words = json.loads(sentiment_config.positive_words) if sentiment_config.positive_words else []
        negative_words = json.loads(sentiment_config.negative_words) if sentiment_config.negative_words else []
        ai_sentiment = analyze_sentiment_local(result.answer or '', '', positive_words, negative_words)
    
    # 更新数据库
    result.ai_sentiment_result = json.dumps(ai_sentiment)
    result.ai_sentiment_updated_at = now_cst()
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': '舆情分析已更新',
        'ai_sentiment': ai_sentiment,
        'updated_at': result.ai_sentiment_updated_at.strftime('%Y-%m-%d %H:%M:%S')
    })


@app.route('/api/results/<int:result_id>', methods=['DELETE'])
@login_required
def delete_result(result_id):
    """删除单条采集结果"""
    result = db.session.get(CollectionResult, result_id)
    if not result:
        return jsonify({'success': False, 'message': '结果不存在'}), 404
    
    task = db.session.get(MonitorTask, result.task_id)
    if not task or task.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权访问'}), 403
    
    db.session.delete(result)
    db.session.commit()
    
    return jsonify({'success': True, 'message': '已删除'})


@app.route('/task/<int:task_id>/results')
@login_required
def view_results(task_id):
    """查看任务结果页面"""
    task = db.session.get(MonitorTask, task_id)
    if not task or task.user_id != current_user.id:
        return redirect(url_for('dashboard'))

    return render_template('results.html', task_id=task_id)


@app.route('/task/<int:task_id>/competitor-analysis')
@login_required
def view_competitor_analysis(task_id):
    """查看竞品分析页面"""
    task = db.session.get(MonitorTask, task_id)
    if not task or task.user_id != current_user.id:
        return redirect(url_for('dashboard'))
    
    return render_template('competitor_analysis.html', task_id=task_id)


@app.route('/api/tasks/<int:task_id>/competitor-analysis', methods=['GET'])
@login_required
def get_competitor_analysis(task_id):
    """获取竞品分析数据"""
    task = db.session.get(MonitorTask, task_id)
    if not task or task.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权访问'}), 403
    
    # 获取竞品品牌列表
    competitor_brands = json.loads(task.competitor_brands) if task.competitor_brands else []
    if not competitor_brands:
        return jsonify({'success': True, 'data': [], 'platforms': []})
    
    # 获取任务配置的平台列表
    task_platforms = json.loads(task.platforms) if task.platforms else []
    
    # 获取日期范围参数
    date_start = request.args.get('date_start')
    date_end = request.args.get('date_end')
    
    # 获取平台筛选参数
    platform_filter = request.args.get('platform')
    
    # 构建查询条件
    query = CollectionResult.query.filter(
        CollectionResult.task_id == task_id
    )
    
    if date_start:
        query = query.filter(CollectionResult.created_at >= date_start)
    if date_end:
        query = query.filter(CollectionResult.created_at <= date_end + ' 23:59:59')
    
    # 如果指定了平台筛选
    if platform_filter:
        query = query.filter(CollectionResult.platform == platform_filter)
    
    results = query.all()
    
    # 统计每个竞品品牌的数据
    brand_stats = {}
    
    for brand in competitor_brands:
        brand_stats[brand] = {
            'brand': brand,
            'mention_count': 0,
            'total_count': 0,
            'rank_sum': 0,
            'rank_count': 0,
            'sentiment_scores': []
        }
    
    all_brands = competitor_brands
    
    for result in results:
        answer = result.answer or ''
        
        for brand in all_brands:
            if brand and answer.find(brand) != -1:
                brand_stats[brand]['mention_count'] += 1
                
                # 尝试提取排名信息
                try:
                    rankings = json.loads(result.rankings) if result.rankings else []
                    if rankings:
                        # 找到该品牌在排名中的位置
                        for rank_info in rankings:
                            name = rank_info.get('name', '')
                            hospital = rank_info.get('hospital', '')
                            if name and brand in name or hospital and brand in hospital:
                                brand_stats[brand]['rank_sum'] += rank_info.get('rank', 1)
                                brand_stats[brand]['rank_count'] += 1
                                break
                except:
                    pass
                
                # 获取情感倾向
                try:
                    sentiment_info = json.loads(result.ai_sentiment_result) if result.ai_sentiment_result else {}
                    sentiment_score = sentiment_info.get('score', 0)
                    brand_stats[brand]['sentiment_scores'].append(sentiment_score)
                except:
                    pass
        
        for brand in all_brands:
            brand_stats[brand]['total_count'] += 1
    
    # 计算最终统计数据
    competitor_data = []
    total_results = len(results) if results else 1
    
    for brand in all_brands:
        if not brand:
            continue
            
        stats = brand_stats[brand]
        mention_rate = round((stats['mention_count'] / stats['total_count']) * 100, 2) if stats['total_count'] > 0 else 0
        avg_rank = round(stats['rank_sum'] / stats['rank_count']) if stats['rank_count'] > 0 else None
        
        # 计算情感倾向
        sentiment = 'neutral'
        if stats['sentiment_scores']:
            avg_sentiment = sum(stats['sentiment_scores']) / len(stats['sentiment_scores'])
            if avg_sentiment > 0.3:
                sentiment = 'positive'
            elif avg_sentiment < -0.3:
                sentiment = 'negative'
        
        competitor_data.append({
            'brand': brand,
            'mention_rate': mention_rate,
            'mention_count': stats['mention_count'],
            'avg_rank': avg_rank,
            'sentiment': sentiment,
            'score': ''
        })
    
    # 按提及率排序
    competitor_data.sort(key=lambda x: x['mention_rate'], reverse=True)
    
    return jsonify({'success': True, 'data': competitor_data, 'platforms': task_platforms})


@app.route('/api/tasks/<int:task_id>/export', methods=['GET'])
@login_required
def export_task_results(task_id):
    """导出任务结果为 Excel，截图直接嵌入单元格"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from flask import send_file
    from PIL import Image as PILImage

    task = db.session.get(MonitorTask, task_id)
    if not task or task.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权访问'}), 403

    platform_filter = request.args.get('platform', '')
    date_start = request.args.get('date_start', '')  # YYYY-MM-DD
    date_end   = request.args.get('date_end', '')    # YYYY-MM-DD

    query = CollectionResult.query.filter_by(task_id=task_id)
    # 跳过空字符串的平台过滤参数
    if platform_filter:
        query = query.filter_by(platform=platform_filter)
    if date_start:
        from datetime import datetime as _dt
        query = query.filter(CollectionResult.created_at >= _dt.strptime(date_start, '%Y-%m-%d'))
    if date_end:
        from datetime import datetime as _dt, timedelta as _td
        query = query.filter(CollectionResult.created_at < _dt.strptime(date_end, '%Y-%m-%d') + _td(days=1))
    results = query.order_by(
        CollectionResult.platform,
        CollectionResult.question,
        CollectionResult.created_at
    ).all()
    if not results:
        return jsonify({'success': False, 'message': '当前筛选条件下暂无可导出的采集结果'}), 404

    platform_names = {
        'doubao': '豆包', 'deepseek': 'DeepSeek', 'kimi': 'Kimi',
        'yuanbao': '元宝', 'wenxin': '文心', 'yiyan': '文心', 'qianwen': '千问',
        'chatgpt': 'ChatGPT'
    }

    wb = Workbook()
    ws = wb.active
    ws.title = '采集结果'

    # ── 样式 ──────────────────────────────────────────────
    header_font  = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill('solid', fgColor='1890FF')
    cell_font    = Font(name='微软雅黑', size=10)
    wrap_align   = Alignment(wrap_text=True, vertical='top')
    center_align = Alignment(horizontal='center', vertical='top')
    thin_border  = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    alt_fill = PatternFill('solid', fgColor='F5F9FF')

    POSITIVE_WORDS = frozenset([
        '好', '不错', '优秀', '棒', '赞', '推荐', '喜欢', '满意', '靠谱', '好评',
        '出色', '卓越', '完美', '理想', '认可', '信赖', '放心', '值得', '划算',
        '实惠', '优质', '高端', '专业', '有效', '神奇', '见效', '吸收', '补水',
        '保湿', '美白', '抗衰', '修复', '温和', '清爽', '不油腻', '无刺激', '抗过敏',
        '口碑好', '效果好', '质量好', '性价比高', '回购', '正品', '安全', '健康',
        '自然', '纯正', '地道', '正宗'
    ])
    NEGATIVE_WORDS = frozenset([
        '差', '不好', '烂', '垃圾', '失望', '后悔', '差评', '糟糕', '骗人', '虚假',
        '没效果', '无用', '无效', '过敏', '刺激', '油腻', '干燥', '紧绷', '爆痘',
        '泛红', '瘙痒', '副作用', '假货', '劣质', '粗糙', '廉价', '智商税', '坑人',
        '受骗', '上当', '不推荐', '避雷', '踩坑', '鸡肋', '不值', '浪费', '昂贵',
        '伤皮肤', '有问题', '不合格', '超标', '有害', '致癌', '激素', '荧光剂', '重金属'
    ])

    def analyze_sentiment(text, keyword=None):
        if not text:
            return '中性'
        ctx = text
        if keyword:
            ki = text.find(keyword)
            if ki != -1:
                start = max(0, ki - 50)
                end = min(len(text), ki + len(keyword) + 50)
                ctx = text[start:end]
        pos = sum(ctx.count(w) for w in POSITIVE_WORDS)
        neg = sum(ctx.count(w) for w in NEGATIVE_WORDS)
        if pos + neg == 0:
            return '中性'
        ratio = (pos - neg) / (pos + neg)
        if ratio > 0.3:
            return '正面'
        elif ratio < -0.3:
            return '负面'
        return '中性'

    # ── 表头（去掉截图路径列，改为截图列） ────────────────
    headers    = ['AI平台', 'AI问题', 'AI答案', '引用参考', '截图', '品牌曝光', '曝光关键词', '舆情倾向', '采集时间']
    col_widths = [12,       30,       60,       50,         30,     10,         20,           10,         20]
    # 截图列索引（1-based）
    IMG_COL = 5

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'

    webapp_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir   = os.path.dirname(webapp_dir)

    for row_idx, r in enumerate(results, 2):
        try:
            # 限制单元格内容长度，Excel 最大支持 32767
            def limit_content(text, limit=30000):
                if not text: return ""
                if len(text) > limit:
                    return text[:limit] + "...(内容过长已截断)"
                return text

            try:
                refs = json.loads(r.references) if r.references else []
            except Exception as e:
                print(f"解析引用失败 row={row_idx}: {e}")
                refs = []

            ref_text = '\n'.join(
                f"{i+1}. {ref.get('title','无标题')}\n   {ref.get('url','')}"
                for i, ref in enumerate(refs)
            ) if refs else ''

            try:
                exposed_kws = json.loads(r.exposed_keywords) if r.exposed_keywords else []
            except Exception as e:
                print(f"解析曝光词失败 row={row_idx}: {e}")
                exposed_kws = []

            # 文字列数据（截图列留空，后面插图片）
            sentiment_result = ''
            if exposed_kws:
                sentiments = [analyze_sentiment(r.answer, kw) for kw in exposed_kws]
                sentiment_items = [f"{kw}({s})" for kw, s in zip(exposed_kws, sentiments)]
                sentiment_result = '、'.join(sentiment_items)

            row_data = [
                platform_names.get(r.platform, r.platform),  # 1
                limit_content(r.question),                    # 2
                limit_content(r.answer),                      # 3
                limit_content(ref_text),                      # 4
                '',                                            # 5 截图（占位）
                '是' if r.has_brand_exposure else '否',        # 6
                '、'.join(exposed_kws),                        # 7
                sentiment_result,                              # 8 舆情倾向
                r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else ''  # 9
            ]

            fill = alt_fill if row_idx % 2 == 0 else None

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = wrap_align if col in (2, 3, 4) else center_align
                if fill:
                    cell.fill = fill

            # ── 嵌入截图 ──────────────────────────────────────
            img_row_height = 20  # 默认行高（无截图时）

            if r.screenshot_path and os.path.isfile(r.screenshot_path):
                try:
                    # 检查文件大小，避免内存溢出（限制 50MB）
                    if os.path.getsize(r.screenshot_path) > 50 * 1024 * 1024:
                        print(f"截图过大跳过 row={row_idx}")
                        ws.cell(row=row_idx, column=IMG_COL, value='截图过大已跳过')
                    else:
                        with PILImage.open(r.screenshot_path) as pil_img:
                            orig_w, orig_h = pil_img.size

                        # 目标显示高度（pt → px：1pt ≈ 1.333px）
                        TARGET_H_PT = 200
                        TARGET_H_PX = int(TARGET_H_PT * 1.333)
                        scale = TARGET_H_PX / orig_h if orig_h > 0 else 1.0
                        display_w = int(orig_w * scale)
                        display_h = TARGET_H_PX

                        # 插入原始图片
                        xl_img = XLImage(r.screenshot_path)
                        xl_img.width  = display_w
                        xl_img.height = display_h

                        col_letter = get_column_letter(IMG_COL)
                        ws.add_image(xl_img, f'{col_letter}{row_idx}')

                        # 列宽自适应显示宽度
                        needed_col_w = display_w / 7 + 2
                        cur_col_w = ws.column_dimensions[get_column_letter(IMG_COL)].width
                        if needed_col_w > cur_col_w:
                            ws.column_dimensions[get_column_letter(IMG_COL)].width = needed_col_w

                        img_row_height = TARGET_H_PT + 4
                except Exception as e:
                    print(f'截图嵌入失败 row={row_idx}: {e}')
                    ws.cell(row=row_idx, column=IMG_COL, value='截图加载失败')

            # 答案行高取答案行数和图片高度的较大值
            answer_lines = len((r.answer or '').split('\n'))
            text_height  = max(20, answer_lines * 14)
            ws.row_dimensions[row_idx].height = min(max(text_height, img_row_height), 400)
            
        except Exception as e:
            print(f"处理行失败 row={row_idx}: {e}")
            ws.cell(row=row_idx, column=1, value=f"错误: 该行导出失败")

    # ── 输出 ──────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    platform_label = platform_names.get(platform_filter, platform_filter) if platform_filter else '全部平台'
    date_label = f'_{date_start}_{date_end}' if date_start and date_end else ''
    filename = f"{task.name}_{platform_label}{date_label}_{task_id}.xlsx"

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/api/tasks/<int:task_id>/export-geo', methods=['GET'])
@login_required
def export_geo_results(task_id):
    """导出GEO效果为Excel，支持按AI平台和时间筛选"""
    import io
    import os
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from flask import send_file
    from PIL import Image as PILImage

    task = db.session.get(MonitorTask, task_id)
    if not task or task.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权访问'}), 403

    # 获取筛选参数
    date_start = request.args.get('date_start', '')  # YYYY-MM-DD
    date_end   = request.args.get('date_end', '')    # YYYY-MM-DD
    platform   = request.args.get('platform', '')    # AI平台筛选（如：doubao, deepseek, kimi等）

    # 调试日志
    print(f"[导出GEO] task_id={task_id}, platform={repr(platform)}, date_start={repr(date_start)}, date_end={repr(date_end)}")

    query = CollectionResult.query.filter_by(task_id=task_id)
    
    # 按平台筛选
    if platform:
        print(f"[导出GEO] 应用平台筛选: {platform}")
        query = query.filter_by(platform=platform)
    
    # 按时间筛选
    if date_start:
        from datetime import datetime as _dt
        query = query.filter(CollectionResult.created_at >= _dt.strptime(date_start, '%Y-%m-%d'))
    if date_end:
        from datetime import datetime as _dt, timedelta as _td
        query = query.filter(CollectionResult.created_at < _dt.strptime(date_end, '%Y-%m-%d') + _td(days=1))
    
    results = query.order_by(
        CollectionResult.platform,
        CollectionResult.question,
        CollectionResult.created_at
    ).all()
    if not results:
        return jsonify({'success': False, 'message': '当前筛选条件下暂无可导出的GEO效果数据'}), 404

    platform_names = {
        'doubao': '豆包', 'deepseek': 'DeepSeek', 'kimi': 'Kimi',
        'yuanbao': '元宝', 'wenxin': '文心', 'yiyan': '文心', 'qianwen': '千问',
        'chatgpt': 'ChatGPT'
    }
    
    brand_keywords = json.loads(task.brand_keywords) if task.brand_keywords else []

    wb = Workbook()
    
    # ==================== Sheet1: GEO效果汇总汇总 ====================
    ws1 = wb.active
    ws1.title = 'GEO效果汇总'
    
    # 样式定义
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1890FF')
    black_header_fill = PatternFill('solid', fgColor='000000')
    cell_font = Font(name='微软雅黑', size=10)
    wrap_align = Alignment(wrap_text=True, vertical='top')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    # 获取所有唯一日期
    all_dates = sorted(set(r.created_at.strftime('%Y-%m-%d') for r in results))
    
    # 第一行：监控问题（合并单元格）
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5 + len(all_dates))
    ws1.cell(row=1, column=1, value='GEO效果')
    ws1.cell(row=1, column=1).font = Font(name='微软雅黑', bold=True, size=12, color='000000')
    ws1.cell(row=1, column=1).alignment = center_align
    ws1.row_dimensions[1].height = 30
    
    # 第二行：表头
    headers = ['序号', 'AI平台', '关键词', 'KPI', '达标天数'] + all_dates
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=header)
        cell.font = Font(name='微软雅黑', bold=True, size=11, color=Color('000000'))
        cell.fill = PatternFill(patternType='solid', fgColor=Color('FFFFFF'))
        cell.alignment = center_align
        cell.border = thin_border
    ws1.row_dimensions[2].height = 28
    
    # 按平台+问题分组统计
    platform_questions = {}
    for r in results:
        platform = r.platform
        question = r.question
        date = r.created_at.strftime('%Y-%m-%d')
        has_exposure = r.has_brand_exposure or any(kw in (r.answer or '') for kw in brand_keywords)
        
        if platform not in platform_questions:
            platform_questions[platform] = {}
        if question not in platform_questions[platform]:
            platform_questions[platform][question] = {
                'dates': set(),
                'exposed_dates': set()
            }
        platform_questions[platform][question]['dates'].add(date)
        if has_exposure:
            platform_questions[platform][question]['exposed_dates'].add(date)
    
    # 填充数据行
    row_idx = 3
    serial_num = 1
    for platform in sorted(platform_questions.keys()):
        for question in sorted(platform_questions[platform].keys()):
            data = platform_questions[platform][question]
            exposed_dates = data['exposed_dates']
            qualified_days = len(exposed_dates)
            
            # 序号
            ws1.cell(row=row_idx, column=1, value=serial_num).alignment = center_align
            # AI平台
            ws1.cell(row=row_idx, column=2, value=platform_names.get(platform, platform)).alignment = center_align
            # 关键词（监控问题）
            ws1.cell(row=row_idx, column=3, value=question).alignment = wrap_align
            # KPI（空值）
            ws1.cell(row=row_idx, column=4, value='').alignment = center_align
            # 达标天数
            ws1.cell(row=row_idx, column=5, value=qualified_days).alignment = center_align
            
            # 日期列：达标显示√，否则空
            for date_idx, date in enumerate(all_dates):
                col = 6 + date_idx
                if date in exposed_dates:
                    ws1.cell(row=row_idx, column=col, value='√').alignment = center_align
                else:
                    ws1.cell(row=row_idx, column=col, value='')
            
            row_idx += 1
            serial_num += 1
    
    # 设置列宽
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 30
    ws1.column_dimensions['D'].width = 8
    ws1.column_dimensions['E'].width = 12
    for i, _ in enumerate(all_dates):
        ws1.column_dimensions[get_column_letter(6 + i)].width = 10
    
    # ==================== Sheet2: GEO效果截图 ====================
    ws2 = wb.create_sheet(title='GEO效果截图')
    
    # 第二行：表头（没有KPI和达标天数）
    headers2 = ['序号', 'AI平台', '关键词'] + all_dates
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=col, value=header)
        cell.font = Font(name='微软雅黑', bold=True, size=11, color=Color('000000'))
        cell.fill = PatternFill(patternType='solid', fgColor=Color('FFFFFF'))
        cell.alignment = center_align
        cell.border = thin_border
    ws2.row_dimensions[2].height = 28
    
    # 按平台+问题分组，获取截图（新逻辑：有截图就收集，同时记录是否有品牌展示）
    platform_question_screenshots = {}
    for r in results:
        platform = r.platform
        question = r.question
        date = r.created_at.strftime('%Y-%m-%d')
        has_exposure = r.has_brand_exposure or any(kw in (r.answer or '') for kw in brand_keywords)
        
        if platform not in platform_question_screenshots:
            platform_question_screenshots[platform] = {}
        if question not in platform_question_screenshots[platform]:
            platform_question_screenshots[platform][question] = {}
        # 新逻辑：有截图就收集，同时记录是否有品牌展示
        if r.screenshot_path:
            platform_question_screenshots[platform][question][date] = {
                'path': r.screenshot_path,
                'has_exposure': has_exposure
            }
    
    # 填充数据行
    row_idx = 3
    serial_num = 1
    webapp_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir = os.path.dirname(webapp_dir)
    
    for platform in sorted(platform_question_screenshots.keys()):
        for question in sorted(platform_question_screenshots[platform].keys()):
            screenshots = platform_question_screenshots[platform][question]
            
            # 序号
            ws2.cell(row=row_idx, column=1, value=serial_num).alignment = center_align
            # AI平台
            ws2.cell(row=row_idx, column=2, value=platform_names.get(platform, platform)).alignment = center_align
            # 关键词（监控问题）
            ws2.cell(row=row_idx, column=3, value=question).alignment = wrap_align
            
            # 设置当前行高度以容纳图片
            ws2.row_dimensions[row_idx].height = 120
            
            # 日期列：有截图就插入，没有品牌展示的单元格边框标红
            for date_idx, date in enumerate(all_dates):
                col = 4 + date_idx
                if date in screenshots:
                    screenshot_info = screenshots[date]
                    screenshot_path = screenshot_info['path']
                    has_exposure = screenshot_info['has_exposure']
                    
                    # 查找实际文件路径
                    full_path = _resolve_screenshot_path(screenshot_path)
                    
                    if full_path:
                        try:
                            img = XLImage(full_path)
                            # 调整图片大小（高度约100px）
                            orig_w, orig_h = img.width, img.height
                            scale = 100 / orig_h if orig_h > 0 else 1.0
                            img.width = int(orig_w * scale)
                            img.height = 100
                            # 设置图片位置偏移，确保在单元格内正确显示
                            img.anchor = f'{get_column_letter(col)}{row_idx}'
                            img.dx = 1000  # 水平偏移
                            img.dy = 1000  # 垂直偏移
                            ws2.add_image(img)
                        except Exception as e:
                            print(f'截图嵌入失败: {e}')
                    
                    # 没有品牌展示的单元格边框标红
                    if not has_exposure:
                        red_border = Border(
                            left=Side(style='thin', color='FF0000'),
                            right=Side(style='thin', color='FF0000'),
                            top=Side(style='thin', color='FF0000'),
                            bottom=Side(style='thin', color='FF0000')
                        )
                        ws2.cell(row=row_idx, column=col).border = red_border
                # 没有截图则留空
            
            row_idx += 1
            serial_num += 1
    
    # 设置列宽
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 30
    for i, _ in enumerate(all_dates):
        ws2.column_dimensions[get_column_letter(4 + i)].width = 30
    
    # ==================== Sheet3+: AI回答（按平台分组） ====================
    # 按平台分组结果
    platform_results = {}
    for r in results:
        platform = r.platform
        if platform not in platform_results:
            platform_results[platform] = []
        platform_results[platform].append(r)
    
    # 为每个平台创建AI回答sheet
    for platform in sorted(platform_results.keys()):
        platform_name = platform_names.get(platform, platform)
        ws = wb.create_sheet(title=f'{platform_name}-AI回答')
        
        # 获取该平台的所有结果和日期
        platform_result_list = platform_results[platform]
        platform_dates = sorted(set(r.created_at.strftime('%Y-%m-%d') for r in platform_result_list))
        
        # 表头：序号、AI平台、关键词、{采集时间}...
        headers = ['序号', 'AI平台', '关键词'] + platform_dates
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(name='微软雅黑', bold=True, size=11, color=Color('FFFFFF'))
            cell.fill = PatternFill(patternType='solid', fgColor=Color('1890FF'))
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[1].height = 28
        
        # 按问题分组
        question_results = {}
        for r in platform_result_list:
            question = r.question
            date = r.created_at.strftime('%Y-%m-%d')
            if question not in question_results:
                question_results[question] = {}
            question_results[question][date] = r
        
        # 填充数据行
        row_idx = 2
        serial_num = 1
        for question in sorted(question_results.keys()):
            date_results = question_results[question]
            
            # 序号
            ws.cell(row=row_idx, column=1, value=serial_num).alignment = center_align
            # AI平台
            ws.cell(row=row_idx, column=2, value=platform_name).alignment = center_align
            # 关键词（监控问题）
            ws.cell(row=row_idx, column=3, value=question).alignment = wrap_align
            
            # 日期列：AI回答内容和引用参考信息的汇总
            for date_idx, date in enumerate(platform_dates):
                col = 4 + date_idx
                if date in date_results:
                    r = date_results[date]
                    # 汇总AI回答内容和引用参考信息
                    content = []
                    if r.answer:
                        content.append(f"AI回答：{r.answer[:500]}")
                    if r.references:
                        # 解析JSON格式的引用参考，提取所有URL
                        try:
                            refs = json.loads(r.references)
                            print(f"[导出GEO] 解析引用参考: 平台={platform}, 问题={question[:20]}, 原始长度={len(r.references)}, 解析后数量={len(refs) if isinstance(refs, list) else 'N/A'}")
                            if isinstance(refs, list) and refs:
                                ref_urls = []
                                for ref in refs:
                                    if isinstance(ref, dict):
                                        url = ref.get('url', '')
                                        title = ref.get('title', '')
                                        if url:
                                            if title:
                                                ref_urls.append(f"{title}: {url}")
                                            else:
                                                ref_urls.append(url)
                                    elif isinstance(ref, str):
                                        ref_urls.append(ref)
                                print(f"[导出GEO] 提取到 {len(ref_urls)} 个引用链接")
                                if ref_urls:
                                    content.append(f"引用参考：\n" + "\n".join(ref_urls))
                                else:
                                    content.append(f"引用参考：{r.references}")
                            else:
                                content.append(f"引用参考：{r.references}")
                        except (json.JSONDecodeError, TypeError) as e:
                            # 如果解析失败，直接使用原始内容
                            print(f"[导出GEO] 引用参考解析失败: {e}")
                            content.append(f"引用参考：{r.references}")
                    cell_value = '\n'.join(content) if content else ''
                    ws.cell(row=row_idx, column=col, value=cell_value).alignment = wrap_align
                else:
                    ws.cell(row=row_idx, column=col, value='')
            
            # 设置行高以容纳更多引用参考链接
            ws.row_dimensions[row_idx].height = 150
            
            row_idx += 1
            serial_num += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        for i, _ in enumerate(platform_dates):
            ws.column_dimensions[get_column_letter(4 + i)].width = 40
    
    # 输出
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    date_label = f'_{date_start}_{date_end}' if date_start and date_end else ''
    filename = f"{task.name}_GEO效果{date_label}_{task_id}.xlsx"
    saved_path = _maybe_save_desktop_download(buf, filename)
    if saved_path:
        return jsonify({
            'success': True,
            'message': '导出成功',
            'filename': os.path.basename(saved_path),
            'path': saved_path
        })
    
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/api/tasks/<int:task_id>/export-screenshots-zip', methods=['GET'])
@login_required
def export_screenshots_zip(task_id):
    """导出GEO长截图为ZIP文件"""
    import io
    import os
    import zipfile
    from flask import send_file
    
    task = db.session.get(MonitorTask, task_id)
    if not task or task.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权访问'}), 403
    
    # 获取筛选参数
    platform_filter = request.args.get('platform', '')
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')
    
    # 查询采集结果
    query = CollectionResult.query.filter(
        CollectionResult.task_id == task_id,
        CollectionResult.screenshot_path.isnot(None)
    )
    
    if platform_filter:
        query = query.filter(CollectionResult.platform == platform_filter)
    
    if date_start:
        query = query.filter(CollectionResult.created_at >= date_start)
    if date_end:
        query = query.filter(CollectionResult.created_at <= date_end + ' 23:59:59')
    
    results = query.all()
    
    if not results:
        return jsonify({'success': False, 'message': '没有找到可导出的截图'}), 404
    
    # 创建ZIP文件
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        webapp_dir = os.path.dirname(os.path.abspath(__file__))
        
        for result in results:
            screenshot_path = result.screenshot_path
            if not screenshot_path:
                continue
            
            # 确定截图文件路径
            screenshot_file = _resolve_screenshot_path(screenshot_path)
            
            if screenshot_file and os.path.exists(screenshot_file):
                # 生成文件名：平台_问题_时间戳.png
                timestamp = result.created_at.strftime('%Y%m%d_%H%M%S')
                platform = result.platform
                question = result.question[:20] if result.question else 'unknown'
                # 清理文件名中的非法字符
                safe_question = ''.join(c for c in question if c not in '<>:"/\\|?*')
                filename = f"{platform}_{safe_question}_{timestamp}.png"
                
                zip_file.write(screenshot_file, filename)
    
    zip_buffer.seek(0)
    
    filename = f"GEO长截图_{task.name}_{now_cst().strftime('%Y%m%d_%H%M%S')}.zip"
    saved_path = _maybe_save_desktop_download(zip_buffer, filename)
    if saved_path:
        return jsonify({
            'success': True,
            'message': '导出成功',
            'filename': os.path.basename(saved_path),
            'path': saved_path
        })
    
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=filename
    )


@app.route('/screenshots/<path:filepath>')
@login_required
def serve_screenshot(filepath):
    """提供截图文件访问
    
    filepath 可能是：
      - web_app/answers/<platform>/...  (web_app 采集)
      - answers/<platform>/...          (main.py 采集，存根目录)
    """
    import os
    from flask import send_file
    
    full_path = _resolve_screenshot_path(filepath)
    if full_path:
        return send_file(full_path, mimetype='image/png')
    
    return jsonify({'error': '截图文件不存在', 'path': filepath}), 404



# 重新采集任务状态存储 { job_id: { status, result_id, error } }
_recollect_jobs = {}


@app.route('/api/tasks/<int:task_id>/recollect', methods=['POST'])
@login_required
def recollect_single(task_id):
    """针对指定问题+平台重新采集，结果覆盖该问题+平台的最新一条记录"""
    task = db.session.get(MonitorTask, task_id)
    if not task or task.user_id != current_user.id:
        return jsonify({'success': False, 'message': '无权访问'}), 403

    data = request.get_json() or {}
    question = data.get('question')
    platform_id = data.get('platform')
    result_id = data.get('result_id')

    if not question or not platform_id:
        return jsonify({'success': False, 'message': '缺少 question 或 platform 参数'}), 400

    import uuid
    job_id = str(uuid.uuid4())
    _recollect_jobs[job_id] = {'status': 'running', 'result_id': None, 'error': None}

    def run_recollect():
        import os
        import sys
        import importlib
        import config as _config
        from playwright.sync_api import sync_playwright
        from browser_utils import launch_browser
        from datetime import datetime, timezone, timedelta

        def now_cst():
            return datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=8))).replace(tzinfo=None)


        with app.app_context():
            t = db.session.get(MonitorTask, task_id)
            brand_keywords = json.loads(t.brand_keywords)
            screenshot_config = json.loads(t.screenshot_config) if t.screenshot_config else {}
            enable_screenshot = screenshot_config.get(platform_id, True)

            WEBAPP_ANSWERS_DIR = answers_dir()
            os.makedirs(WEBAPP_ANSWERS_DIR, exist_ok=True)
            _original = _config.OUTPUT_DIR
            _config.OUTPUT_DIR = WEBAPP_ANSWERS_DIR

            try:
                platform_module = importlib.import_module(f'platforms.{platform_id}')
                user_data_dir = get_profile_dir(platform_id, t.user_id)

                def _collect_pngs(d):
                    found = set()
                    if os.path.exists(d):
                        for root, _, files in os.walk(d):
                            for f in files:
                                if f.lower().endswith('.png'):
                                    found.add(os.path.join(root, f))
                    return found

                with sync_playwright() as p:
                    context, browser = launch_browser(p, headless=False, user_data_dir=user_data_dir)
                    page = context.pages[0] if context.pages else context.new_page()
                    try:
                        pngs_before = _collect_pngs(WEBAPP_ANSWERS_DIR)
                        # 修复：传递 brand_keywords 和 enable_screenshot，否则长截图不会画红框
                        answer, references = platform_module.query(
                            page, 
                            question, 
                            brand_keywords=brand_keywords, 
                            enable_screenshot=enable_screenshot
                        )

                        screenshot_path = None
                        if enable_screenshot:
                            import time as _t
                            _t.sleep(0.5)
                            new_pngs = _collect_pngs(WEBAPP_ANSWERS_DIR) - pngs_before
                            if new_pngs:
                                screenshot_path = max(new_pngs, key=os.path.getmtime)

                        has_exposure = any(kw in (answer or '') for kw in brand_keywords)
                        exposed_keywords = [kw for kw in brand_keywords if kw in (answer or '')]

                        target = None
                        if result_id:
                            target = db.session.get(CollectionResult, result_id)
                        if not target:
                            target = (CollectionResult.query
                                      .filter_by(task_id=task_id, question=question, platform=platform_id)
                                      .order_by(CollectionResult.created_at.desc())
                                      .first())

                        if target:
                            target.answer = answer
                            target.references = json.dumps(references, ensure_ascii=False)
                            target.screenshot_path = screenshot_path
                            target.has_brand_exposure = has_exposure
                            target.exposed_keywords = json.dumps(exposed_keywords, ensure_ascii=False)
                            target.created_at = now_cst()
                        else:
                            target = CollectionResult(
                                task_id=task_id,
                                question=question,
                                platform=platform_id,
                                answer=answer,
                                references=json.dumps(references, ensure_ascii=False),
                                screenshot_path=screenshot_path,
                                has_brand_exposure=has_exposure,
                                exposed_keywords=json.dumps(exposed_keywords, ensure_ascii=False)
                            )
                            db.session.add(target)

                        db.session.commit()
                        
                        # 重新采集后，重新计算智慧舆情
                        if t.sentiment_config_id:
                            sentiment_config = SentimentConfig.query.filter_by(id=t.sentiment_config_id, user_id=current_user.id).first()
                            if sentiment_config:
                                # 分析智慧舆情
                                ai_sentiment = None
                                if sentiment_config.enable_ai_sentiment and sentiment_config.ai_api_url and sentiment_config.ai_api_key:
                                    ai_sentiment = analyze_sentiment_ai(target.answer or '', '', sentiment_config)
                                    # 如果智能舆情分析失败，降级到普通舆情分析
                                    if ai_sentiment.get('error') or '分析失败' in ai_sentiment.get('reason', ''):
                                        positive_words = json.loads(sentiment_config.positive_words) if sentiment_config.positive_words else []
                                        negative_words = json.loads(sentiment_config.negative_words) if sentiment_config.negative_words else []
                                        local_sentiment = analyze_sentiment_local(target.answer or '', '', positive_words, negative_words)
                                        ai_sentiment = {**local_sentiment, 'reason': f"智能分析失败，已使用关键词分析: {local_sentiment.get('label')}"}
                                elif sentiment_config.positive_words or sentiment_config.negative_words:
                                    # 没有启用智能舆情但有关键词配置
                                    positive_words = json.loads(sentiment_config.positive_words) if sentiment_config.positive_words else []
                                    negative_words = json.loads(sentiment_config.negative_words) if sentiment_config.negative_words else []
                                    ai_sentiment = analyze_sentiment_local(target.answer or '', '', positive_words, negative_words)
                                
                                if ai_sentiment:
                                    target.ai_sentiment_result = json.dumps(ai_sentiment)
                                    target.ai_sentiment_updated_at = now_cst()
                                    db.session.commit()
                        
                        _recollect_jobs[job_id] = {
                            'status': 'done',
                            'result_id': target.id,
                            'result': target.to_dict(),
                            'error': None
                        }
                        print(f'重新采集完成（已覆盖 id={target.id}）: {platform_id} / {question}')
                    finally:
                        context.close()
                        if browser:
                            browser.close()
            except Exception as e:
                import traceback; traceback.print_exc()
                _recollect_jobs[job_id] = {'status': 'error', 'result_id': None, 'error': str(e)}
            finally:
                _config.OUTPUT_DIR = _original

    thread = threading.Thread(target=run_recollect, daemon=True)
    thread.start()

    return jsonify({'success': True, 'job_id': job_id, 'message': f'已开始重新采集：{platform_id} / {question}'})


@app.route('/api/recollect-status/<job_id>', methods=['GET'])
@login_required
def recollect_status(job_id):
    """查询重新采集任务状态"""
    job = _recollect_jobs.get(job_id)
    if not job:
        return jsonify({'success': False, 'message': '任务不存在'}), 404
    return jsonify({'success': True, **job})


@app.route('/api/check-login/<platform>', methods=['GET'])
@login_required
def check_platform_login_status(platform):
    """检测指定平台的登录状态"""
    if not LOGIN_CHECKER_AVAILABLE:
        return jsonify({
            'success': False,
            'message': '登录检测功能不可用'
        }), 503

    try:
        logger.info(f"[LoginCheck] 开始检测平台登录状态: {platform}")
        # 在后台线程中执行检测（避免阻塞）
        result = check_platform_login(platform, current_user.id)
        
        logger.info(f"[LoginCheck] 平台登录状态检测完成: {result['name']} - {'已登录' if result['is_logged_in'] else '未登录'}")
        _set_cached_login_status(current_user.id, result['platform'], result['is_logged_in'], result.get('error'))
        cached = _normalize_cached_login_status(_get_cached_login_status(current_user.id).get(result['platform']))
        return jsonify({
            'success': True,
            'platform': result['platform'],
            'name': result['name'],
            'is_logged_in': result['is_logged_in'],
            'checked_at': cached.get('checked_at'),
            'error': result['error']
        })
    except Exception as e:
        logger.error(f"[LoginCheck] 平台登录状态检测失败: {platform} - {str(e)}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/check-login/all', methods=['GET'])
@login_required
def check_all_platforms_status():
    """检测所有平台的登录状态"""
    if not LOGIN_CHECKER_AVAILABLE:
        return jsonify({
            'success': False,
            'message': '登录检测功能不可用'
        }), 503

    if app.config.get('DESKTOP_MODE'):
        cached_status = _get_cached_login_status(current_user.id)
        platforms = Config.SUPPORTED_PLATFORMS
        return jsonify({
            'success': True,
            'platforms': [
                {
                    'platform': p['id'],
                    'name': p['name'],
                    'is_logged_in': _normalize_cached_login_status(cached_status.get(p['id'], False))['is_logged_in'],
                    'checked_at': _normalize_cached_login_status(cached_status.get(p['id'], False))['checked_at'],
                    'error': None if _normalize_cached_login_status(cached_status.get(p['id'], False))['is_logged_in'] else '单机版请按平台逐个点击“重新检测”，避免一次性打开多个浏览器'
                }
                for p in platforms
            ],
            'message': '单机版已跳过批量检测'
        })
    
    try:
        logger.info("[LoginCheck] 开始检测所有平台登录状态")
        results = check_all_platforms(current_user.id)
        
        logger.info("[LoginCheck] 所有平台登录状态检测完成")
        return jsonify({
            'success': True,
            'platforms': results
        })
    except Exception as e:
        logger.error(f"[LoginCheck] 批量检测失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/login-status-cache', methods=['GET', 'POST'])
@login_required
def get_login_status_cache():
    """返回当前用户上次检测得到的平台登录状态，用于页面快速恢复显示。"""
    if request.method == 'POST':
        data = request.get_json() or {}
        incoming_status = data.get('status', {})
        allowed_platforms = {p['id'] for p in Config.SUPPORTED_PLATFORMS}
        if not isinstance(incoming_status, dict):
            return jsonify({'success': False, 'message': '状态格式无效'}), 400

        for platform, is_logged_in in incoming_status.items():
            if platform in allowed_platforms:
                normalized = _normalize_cached_login_status(is_logged_in)
                _set_cached_login_status(
                    current_user.id,
                    platform,
                    normalized['is_logged_in'],
                    normalized.get('error')
                )

        return jsonify({'success': True})

    cached_status = _get_cached_login_status(current_user.id)
    normalized_status = {
        platform: _normalize_cached_login_status(value)
        for platform, value in cached_status.items()
    }
    return jsonify({
        'success': True,
        'status': normalized_status,
        'flat_status': {
            platform: value['is_logged_in']
            for platform, value in normalized_status.items()
        }
    })


@app.route('/api/login/<platform>', methods=['POST'])
@login_required
def open_platform_login(platform):
    """打开浏览器窗口让用户登录指定平台"""
    print(f"\n[登录API] 收到登录请求: platform={platform}")
    print(f"[登录API] LOGIN_HELPER_AVAILABLE = {LOGIN_HELPER_AVAILABLE}")
    
    if not LOGIN_HELPER_AVAILABLE:
        print("[登录API] 错误: LOGIN_HELPER_AVAILABLE = False")
        return jsonify({
            'success': False,
            'message': '浏览器登录功能不可用'
        }), 503
    
    try:
        # 获取等待时间参数（默认60秒）
        data = request.get_json() or {}
        wait_time = data.get('wait_time', 60)
        user_id = current_user.id
        
        # 在后台线程中打开浏览器
        def open_browser_thread():
            try:
                print(f"[登录线程] 开始打开浏览器: user_id={user_id}, platform={platform}")
                result = open_login_browser(platform, wait_time, user_id=user_id)
                print(f"[登录线程] 登录浏览器结果: {result}")
            except Exception as e:
                print(f"[登录线程] 打开登录浏览器失败: {type(e).__name__}: {e}")
                import traceback
                traceback.print_exc()
        
        thread = threading.Thread(target=open_browser_thread)
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'success': True,
            'message': f'正在打开浏览器窗口，请在浏览器中完成登录操作',
            'platform': platform,
            'wait_time': wait_time
        })
    except Exception as e:
        print(f"[登录API] 异常: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/login/<platform>', methods=['DELETE'])
@login_required
def clear_platform_login(platform):
    """清除当前用户指定平台的浏览器登录态/Cookie"""
    try:
        cleared, profile_dir = clear_profile_dir(platform, current_user.id)
        logger.info(
            "[LoginClear] user_id=%s platform=%s cleared=%s profile=%s",
            current_user.id,
            platform,
            cleared,
            profile_dir,
        )
        _set_cached_login_status(current_user.id, platform, False, None)
        return jsonify({
            'success': True,
            'platform': platform,
            'cleared': cleared,
            'message': '登录状态已清除' if cleared else '该平台暂无已保存的登录状态'
        })
    except Exception as e:
        logger.error(f"[LoginClear] 清除平台登录状态失败: {platform} - {str(e)}")
        return jsonify({
            'success': False,
            'message': f'清除登录状态失败: {str(e)}'
        }), 500


# ==================== GEO稿件管理 ====================

@app.route('/api/geo-manuscripts', methods=['GET'])
@login_required
def get_geo_manuscripts():
    """获取GEO稿件列表"""
    task_id = request.args.get('task_id', type=int)
    
    query = GeoManuscript.query.filter_by(user_id=current_user.id)
    if task_id:
        query = query.filter_by(task_id=task_id)
        
    manuscripts = query.order_by(GeoManuscript.created_at.desc()).all()
    return jsonify({
        'success': True,
        'manuscripts': [m.to_dict() for m in manuscripts]
    })

@app.route('/api/geo-manuscripts', methods=['POST'])
@login_required
def add_geo_manuscript():
    """添加或更新GEO稿件（支持批量添加）"""
    try:
        data = request.get_json()
        m_id = data.get('id')
        
        # 检查是否为批量添加
        urls = data.get('urls')
        if urls and isinstance(urls, list) and len(urls) > 0:
            # 批量添加模式
            manuscripts = []
            
            # 获取任务ID列表
            task_ids = data.get('task_ids', [])
            if not isinstance(task_ids, list):
                task_ids = []
            
            # 转换为JSON字符串
            task_ids_json = json.dumps(task_ids) if task_ids else None
            
            # 兼容旧的 task_id 字段
            task_id = data.get('task_id')
            if task_id and not task_ids:
                task_id = int(task_id) if isinstance(task_id, str) else task_id
                task_ids = [task_id]
                task_ids_json = json.dumps(task_ids)
            
            for url in urls:
                if url.strip():
                    m = GeoManuscript(
                        user_id=current_user.id,
                        title=data.get('title'),
                        url=url.strip(),
                        task_id=task_ids[0] if task_ids else None,  # 兼容旧字段
                        task_ids=task_ids_json
                    )
                    manuscripts.append(m)
            
            if manuscripts:
                db.session.add_all(manuscripts)
                db.session.commit()
            
            return jsonify({'success': True, 'count': len(manuscripts)})
        
        if m_id:
            # 更新
            m = GeoManuscript.query.filter_by(id=m_id, user_id=current_user.id).first_or_404()
            m.title = data.get('title', m.title)
            m.url = data.get('url', m.url)
            
            # 处理任务关联
            task_ids = data.get('task_ids', [])
            if not isinstance(task_ids, list):
                task_ids = []
            
            # 转换为JSON字符串
            task_ids_json = json.dumps(task_ids) if task_ids else None
            
            # 兼容旧的 task_id 字段
            task_id = data.get('task_id')
            if task_id and not task_ids:
                task_id = int(task_id) if isinstance(task_id, str) else task_id
                task_ids = [task_id]
                task_ids_json = json.dumps(task_ids)
            
            m.task_id = task_ids[0] if task_ids else None  # 兼容旧字段
            m.task_ids = task_ids_json
        else:
            # 新增单个
            # 处理任务关联
            task_ids = data.get('task_ids', [])
            if not isinstance(task_ids, list):
                task_ids = []
            
            # 转换为JSON字符串
            task_ids_json = json.dumps(task_ids) if task_ids else None
            
            # 兼容旧的 task_id 字段
            task_id = data.get('task_id')
            if task_id and not task_ids:
                task_id = int(task_id) if isinstance(task_id, str) else task_id
                task_ids = [task_id]
                task_ids_json = json.dumps(task_ids)
            
            m = GeoManuscript(
                user_id=current_user.id,
                title=data.get('title'),
                url=data.get('url'),
                task_id=task_ids[0] if task_ids else None,  # 兼容旧字段
                task_ids=task_ids_json
            )
            db.session.add(m)
            
        db.session.commit()
        return jsonify({'success': True, 'manuscript': m.to_dict()})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/geo-manuscripts/<int:m_id>', methods=['DELETE'])
@login_required
def delete_geo_manuscript(m_id):
    """删除GEO稿件"""
    m = GeoManuscript.query.filter_by(id=m_id, user_id=current_user.id).first_or_404()
    db.session.delete(m)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/analysis/geo-coverage', methods=['GET'])
@login_required
def get_geo_coverage_analysis():
    """获取GEO稿件覆盖率分析（是否被引用）"""
    task_id = request.args.get('task_id', type=int)
    platform = request.args.get('platform', '')
    date_str = request.args.get('date', '') # YYYY-MM-DD
    
    # 1. 获取所有相关的 GEO 稿件
    m_query = GeoManuscript.query.filter_by(user_id=current_user.id)
    if task_id:
        m_query = m_query.filter_by(task_id=task_id)
    manuscripts = m_query.all()
    
    # 2. 获取相关任务的所有采集结果
    query = CollectionResult.query.join(MonitorTask).filter(MonitorTask.user_id == current_user.id)
    
    if task_id:
        query = query.filter(CollectionResult.task_id == task_id)
    if platform:
        query = query.filter(CollectionResult.platform == platform)
    if date_str:
        try:
            start_dt = datetime.strptime(date_str, '%Y-%m-%d')
            end_dt = start_dt + timedelta(days=1)
            query = query.filter(CollectionResult.created_at >= start_dt, CollectionResult.created_at < end_dt)
        except ValueError:
            pass
            
    # 按时间倒序，以便在匹配时优先匹配最新的结果
    results = query.order_by(CollectionResult.created_at.desc()).all()
    
    # 3. 统计逻辑
    coverage_data = _analyze_geo_coverage(manuscripts, results)
    
    return jsonify({
        'success': True,
        'data': coverage_data
    })


def _analyze_geo_coverage(manuscripts, results):
    """分析GEO稿件覆盖率（提取为公共函数供导出使用）"""
    coverage_data = []
    
    import re

    def get_core_url(url):
        """提取 URL 的核心部分用于对比"""
        if not url: return ""
        url = url.strip().lower()
        if '://' in url: url = url.split('://', 1)[1]
        # 移除查询参数和锚点
        url = url.split('?')[0].split('#')[0]
        # 移除末尾斜杠
        if url.endswith('/'): url = url[:-1]
        # 移除 www. 前缀
        if url.startswith('www.'): url = url[4:]
        return url

    def extract_ids(url_core):
        """从核心 URL 中提取可能的 ID（连续4位以上数字）"""
        return [s for s in re.split(r'[^0-9]', url_core) if len(s) >= 4]

    def get_main_domain(url_core):
        """提取域名的主干部分（如 toutiao.com）"""
        domain = url_core.split('/')[0]
        parts = domain.split('.')
        if len(parts) >= 2:
            # 处理 .com.cn 等双后缀
            special_suffixes = ['com.cn', 'net.cn', 'gov.cn', 'org.cn', 'edu.cn']
            if '.'.join(parts[-2:]) in special_suffixes and len(parts) >= 3:
                return '.'.join(parts[-3:])
            return '.'.join(parts[-2:])
        return domain

    def get_task_ids_for_manuscript(m):
        """获取稿件关联的所有任务ID（支持新旧字段）"""
        task_ids = []
        
        # 优先使用新字段 task_ids（JSON数组）
        if m.task_ids:
            try:
                task_ids = json.loads(m.task_ids)
                if not isinstance(task_ids, list):
                    task_ids = []
            except:
                task_ids = []
        
        # 如果新字段为空，使用旧字段 task_id
        if not task_ids and m.task_id:
            task_ids = [m.task_id]
        
        return task_ids

    for m in manuscripts:
        m_core = get_core_url(m.url)
        if not m_core: continue
        
        m_ids = extract_ids(m_core)
        m_main_domain = get_main_domain(m_core)
        
        # 获取稿件关联的所有任务ID
        m_task_ids = get_task_ids_for_manuscript(m)
        
        cited_in = []
        for r in results:
            # 稿件关联了具体任务时，严格匹配该任务的结果
            if m_task_ids and r.task_id not in m_task_ids:
                continue
                
            refs = json.loads(r.references) if r.references else []
            for ref in refs:
                ref_url_orig = ref.get('url', '')
                ref_core = get_core_url(ref_url_orig)
                if not ref_core: continue
                
                is_match = False
                # 策略1：核心部分互相包含（最常用，处理子域名或简单路径差异）
                if m_core in ref_core or ref_core in m_core:
                    is_match = True
                # 策略2：如果包含长数字 ID，且 ID 匹配，且主域名一致（处理 article vs group 等路径变体）
                elif m_ids:
                    ref_ids = extract_ids(ref_core)
                    common_ids = set(m_ids) & set(ref_ids)
                    if common_ids:
                        ref_main_domain = get_main_domain(ref_core)
                        if m_main_domain == ref_main_domain:
                            is_match = True
                
                if is_match:
                    cited_in.append({
                        'question': r.question,
                        'platform': r.platform,
                        'ref_title': ref.get('title', ''),
                        'ref_url': ref_url_orig,
                        'created_at': r.created_at.strftime('%Y-%m-%d %H:%M')
                    })
                    break # 一个采集结果中命中一次该稿件即可
        
        # 获取任务名称（支持多个任务）
        task_names = []
        for tid in m_task_ids:
            task = db.session.get(MonitorTask, tid)
            if task:
                task_names.append(task.name)
        
        # 如果没有关联任务，显示"未关联"
        task_name_display = ', '.join(task_names) if task_names else "未关联"
        
        coverage_data.append({
            'id': m.id,
            'title': m.title,
            'url': m.url,
            'task_id': m.task_id,
            'task_ids': m_task_ids,
            'task_name': task_name_display,
            'is_cited': len(cited_in) > 0,
            'cited_count': len(cited_in),
            'details': cited_in
        })
    
    return coverage_data


@app.route('/api/analysis/geo-coverage/export', methods=['GET'])
@login_required
def export_geo_coverage_analysis():
    """导出GEO稿件被引用分析结果为Excel"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file
    
    task_id = request.args.get('task_id', type=int)
    platform = request.args.get('platform', '')
    date_str = request.args.get('date', '')
    
    # 1. 获取所有相关的 GEO 稿件
    m_query = GeoManuscript.query.filter_by(user_id=current_user.id)
    if task_id:
        m_query = m_query.filter_by(task_id=task_id)
    manuscripts = m_query.all()
    
    # 2. 获取相关任务的所有采集结果
    query = CollectionResult.query.join(MonitorTask).filter(MonitorTask.user_id == current_user.id)
    
    if task_id:
        query = query.filter(CollectionResult.task_id == task_id)
    if platform:
        query = query.filter(CollectionResult.platform == platform)
    if date_str:
        try:
            start_dt = datetime.strptime(date_str, '%Y-%m-%d')
            end_dt = start_dt + timedelta(days=1)
            query = query.filter(CollectionResult.created_at >= start_dt, CollectionResult.created_at < end_dt)
        except ValueError:
            pass
            
    results = query.order_by(CollectionResult.created_at.desc()).all()
    
    # 3. 分析数据
    coverage_data = _analyze_geo_coverage(manuscripts, results)
    
    # 4. 创建Excel
    wb = Workbook()
    
    # Sheet1: 汇总表
    ws1 = wb.active
    ws1.title = 'GEO稿件被引用分析'
    
    # 表头样式
    header_font = Font(bold=True, color='000000')
    header_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 表头
    headers = ['序号', '稿件备注', '目标URL/特征', '关联任务', '引用状态', '引用次数']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 填充数据
    platform_name_map = {
        'doubao': '豆包',
        'wenxin': '文心一言',
        'yiyan': '文心一言',
        'qianwen': '千问',
        'yuanbao': '元宝',
        'kimi': 'Kimi',
        'deepseek': 'DeepSeek',
        'chatgpt': 'ChatGPT'
    }
    
    for row, item in enumerate(coverage_data, 2):
        ws1.cell(row=row, column=1, value=row-1).border = thin_border
        ws1.cell(row=row, column=2, value=item['title'] or '-').border = thin_border
        ws1.cell(row=row, column=3, value=item['url']).border = thin_border
        ws1.cell(row=row, column=4, value=item['task_name']).border = thin_border
        ws1.cell(row=row, column=5, value='已引用' if item['is_cited'] else '未引用').border = thin_border
        ws1.cell(row=row, column=6, value=item['cited_count']).border = thin_border
    
    # 设置列宽
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 50
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 12
    
    # Sheet2: 引用来源详情表
    ws2 = wb.create_sheet(title='引用来源详情')
    
    # 表头
    detail_headers = ['序号', '稿件备注', '稿件URL', '关联任务', '引用问题', 'AI平台', '引用来源标题', '引用来源URL', '引用时间']
    for col, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 填充详情数据
    detail_row = 2
    for item in coverage_data:
        if item['details']:
            for detail in item['details']:
                ws2.cell(row=detail_row, column=1, value=detail_row-1).border = thin_border
                ws2.cell(row=detail_row, column=2, value=item['title'] or '-').border = thin_border
                ws2.cell(row=detail_row, column=3, value=item['url']).border = thin_border
                ws2.cell(row=detail_row, column=4, value=item['task_name']).border = thin_border
                ws2.cell(row=detail_row, column=5, value=detail['question'] or '-').border = thin_border
                ws2.cell(row=detail_row, column=6, value=platform_name_map.get(detail['platform'], detail['platform'])).border = thin_border
                ws2.cell(row=detail_row, column=7, value=detail['ref_title'] or '-').border = thin_border
                ws2.cell(row=detail_row, column=8, value=detail['ref_url']).border = thin_border
                ws2.cell(row=detail_row, column=9, value=detail['created_at']).border = thin_border
                detail_row += 1
    
    # 设置列宽
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 35
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 30
    ws2.column_dimensions['F'].width = 12
    ws2.column_dimensions['G'].width = 35
    ws2.column_dimensions['H'].width = 50
    ws2.column_dimensions['I'].width = 18
    
    # 输出
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    filename = f"GEO稿件被引用分析_{date_str or '全部'}_{platform or '全部平台'}.xlsx"
    
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ==================== 浏览器设置 ====================

@app.route('/api/browser/config', methods=['GET'])
@login_required
def get_browser_config():
    import sys
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from browser_config import load_browser_config, get_browser_candidates
    from browser_utils import find_browser
    
    config = load_browser_config()
    current_path = ''
    browser_type = ''
    
    try:
        browser_type, current_path = find_browser()
    except:
        pass
    
    return jsonify({
        'success': True,
        'config': {
            'current_browser_path': current_path,
            'current_browser_type': browser_type,
            'configured_path': config.get('browser_path', ''),
            'candidates': get_browser_candidates()
        }
    })

@app.route('/api/browser/config', methods=['POST'])
@login_required
def save_browser_config():
    try:
        data = request.get_json()
        browser_path = data.get('browser_path', '').strip()
        
        import sys
        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from browser_config import save_browser_config
        
        save_browser_config(browser_path)
        
        return jsonify({
            'success': True,
            'message': '浏览器配置已保存'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


# ==================== 辅助函数 ====================

def calculate_exposure_stats(results, brand_keywords):
    """计算品牌词曝光率统计"""
    stats = {
        'total_questions': 0,
        'total_answers': len(results),
        'exposure_count': 0,
        'exposure_rate': 0,
        'platform_stats': {},
        'keyword_stats': {}
    }
    
    # 统计每个平台的曝光情况
    for result in results:
        platform = result.platform
        answer = result.answer or ''
        
        if platform not in stats['platform_stats']:
            stats['platform_stats'][platform] = {
                'total': 0,
                'exposed': 0,
                'rate': 0
        }
        
        stats['platform_stats'][platform]['total'] += 1
        
        # 检查品牌词是否出现在答案中
        has_exposure = False
        for keyword in brand_keywords:
            if keyword in answer:
                has_exposure = True
                
                # 统计每个关键词的曝光次数
                if keyword not in stats['keyword_stats']:
                    stats['keyword_stats'][keyword] = 0
                stats['keyword_stats'][keyword] += 1
        
        if has_exposure:
            stats['exposure_count'] += 1
            stats['platform_stats'][platform]['exposed'] += 1
    
    # 计算曝光率
    if stats['total_answers'] > 0:
        stats['exposure_rate'] = round(stats['exposure_count'] / stats['total_answers'] * 100, 2)
    
    # 计算每个平台的曝光率
    for platform in stats['platform_stats']:
        total = stats['platform_stats'][platform]['total']
        exposed = stats['platform_stats'][platform]['exposed']
        if total > 0:
            stats['platform_stats'][platform]['rate'] = round(exposed / total * 100, 2)
    
    return stats


@app.route('/api/analysis/references', methods=['GET'])
@login_required
def get_reference_analysis():
    """获取引用参考分析数据"""
    task_id = request.args.get('task_id', type=int)
    platform = request.args.get('platform', '')
    date_start = request.args.get('date_start', '')  # YYYY-MM-DD
    date_end = request.args.get('date_end', '')      # YYYY-MM-DD
    deduplicate = request.args.get('deduplicate', 'true').lower() == 'true'

    # 基础查询：仅限当前用户
    query = CollectionResult.query.join(MonitorTask).filter(MonitorTask.user_id == current_user.id)

    if task_id:
        query = query.filter(CollectionResult.task_id == task_id)
    if platform:
        query = query.filter(CollectionResult.platform == platform)
    
    if date_start:
        try:
            start_dt = datetime.strptime(date_start, '%Y-%m-%d')
            query = query.filter(CollectionResult.created_at >= start_dt)
        except ValueError:
            pass
            
    if date_end:
        try:
            end_dt = datetime.strptime(date_end, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(CollectionResult.created_at < end_dt)
        except ValueError:
            pass

    results = query.all()

    # 分析引用
    media_counts = {}
    seen_urls = set()
    
    from urllib.parse import urlparse

    def get_full_domain(url):
        if not url or not url.startswith('http'):
            return "未知来源"
        try:
            domain = urlparse(url).netloc.lower()
            if ':' in domain:
                domain = domain.split(':')[0]
            if domain.startswith('www.'):
                domain = domain[4:]
            return domain
        except:
            return "未知来源"

    def get_top_domain(domain):
        if domain == "未知来源":
            return domain
        try:
            parts = domain.split('.')
            if len(parts) <= 2:
                return domain
            
            # 处理常见双后缀（如 .com.cn, .gov.cn）
            special_suffixes = ['com.cn', 'net.cn', 'gov.cn', 'org.cn', 'edu.cn', 'bj.cn', 'sh.cn']
            last_two = '.'.join(parts[-2:])
            if last_two in special_suffixes and len(parts) >= 3:
                return '.'.join(parts[-3:])
            
            # 默认返回一级域名（如 news.baidu.com -> baidu.com）
            return '.'.join(parts[-2:])
        except:
            return domain

    # 分析引用
    full_media_counts = {}
    top_media_counts = {}
    seen_urls = set()
    
    for r in results:
        refs = json.loads(r.references) if r.references else []
        for ref in refs:
            url = ref.get('url', '')
            
            # 排重逻辑
            if deduplicate:
                if url and url in seen_urls:
                    continue
                if url:
                    seen_urls.add(url)
            
            if url:
                full_domain = get_full_domain(url)
                top_domain = get_top_domain(full_domain)
                
                full_media_counts[full_domain] = full_media_counts.get(full_domain, 0) + 1
                top_media_counts[top_domain] = top_media_counts.get(top_domain, 0) + 1
            else:
                full_media_counts["未知来源"] = full_media_counts.get("未知来源", 0) + 1
                top_media_counts["未知来源"] = top_media_counts.get("未知来源", 0) + 1

    # 排序并转换为列表
    sorted_full_media = sorted(
        [{'name': k, 'count': v} for k, v in full_media_counts.items()],
        key=lambda x: x['count'],
        reverse=True
    )
    
    sorted_top_media = sorted(
        [{'name': k, 'count': v} for k, v in top_media_counts.items()],
        key=lambda x: x['count'],
        reverse=True
    )

    return jsonify({
        'success': True,
        'full_data': sorted_full_media,
        'top_data': sorted_top_media,
        'total_results': len(results),
        'total_references': sum(full_media_counts.values())
    })


@app.route('/api/analysis/domains', methods=['GET'])
@login_required
def get_reference_domains():
    """获取引用来源域名列表，用于走势图筛选。"""
    task_id = request.args.get('task_id', type=int)
    platform = request.args.get('platform', '')
    level = request.args.get('level', 'top')
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')

    query = CollectionResult.query.join(MonitorTask).filter(MonitorTask.user_id == current_user.id)
    if task_id:
        query = query.filter(CollectionResult.task_id == task_id)
    if platform:
        query = query.filter(CollectionResult.platform == platform)

    if date_start:
        try:
            query = query.filter(CollectionResult.created_at >= datetime.strptime(date_start, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_end:
        try:
            query = query.filter(CollectionResult.created_at < datetime.strptime(date_end, '%Y-%m-%d') + timedelta(days=1))
        except ValueError:
            pass

    from urllib.parse import urlparse

    def get_full_domain(url):
        if not url or not url.startswith('http'):
            return None
        try:
            domain = urlparse(url).netloc.lower()
            if ':' in domain:
                domain = domain.split(':')[0]
            if domain.startswith('www.'):
                domain = domain[4:]
            return domain
        except Exception:
            return None

    def get_top_domain(domain):
        if not domain:
            return None
        try:
            parts = domain.split('.')
            if len(parts) <= 2:
                return domain
            special_suffixes = ['com.cn', 'net.cn', 'gov.cn', 'org.cn', 'edu.cn', 'bj.cn', 'sh.cn']
            last_two = '.'.join(parts[-2:])
            if last_two in special_suffixes and len(parts) >= 3:
                return '.'.join(parts[-3:])
            return '.'.join(parts[-2:])
        except Exception:
            return domain

    counts = {}
    for result in query.all():
        refs = json.loads(result.references) if result.references else []
        for ref in refs:
            full_domain = get_full_domain(ref.get('url', ''))
            domain = get_top_domain(full_domain) if level == 'top' else full_domain
            if domain:
                counts[domain] = counts.get(domain, 0) + 1

    domains = [
        name
        for name, _ in sorted(counts.items(), key=lambda item: item[1], reverse=True)
    ]
    return jsonify({
        'success': True,
        'domains': domains,
        'counts': counts
    })


@app.route('/api/analysis/reference-trends', methods=['GET'])
@login_required
def get_reference_trends():
    """获取引用参考走势数据"""
    task_id = request.args.get('task_id', type=int)
    platform = request.args.get('platform', '')
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')
    top_n = request.args.get('top_n', 10, type=int)
    level = request.args.get('level', 'top') # top 或 full
    # 获取指定的域名列表
    selected_domains = request.args.getlist('domains[]') or request.args.getlist('domains')

    # 默认近一年
    if not date_start:
        date_start = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
    if not date_end:
        date_end = datetime.now().strftime('%Y-%m-%d')

    # 基础查询
    query = CollectionResult.query.join(MonitorTask).filter(MonitorTask.user_id == current_user.id)

    if task_id:
        query = query.filter(CollectionResult.task_id == task_id)
    if platform:
        query = query.filter(CollectionResult.platform == platform)
    
    try:
        start_dt = datetime.strptime(date_start, '%Y-%m-%d')
        query = query.filter(CollectionResult.created_at >= start_dt)
        end_dt = datetime.strptime(date_end, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(CollectionResult.created_at < end_dt)
    except ValueError:
        pass

    results = query.order_by(CollectionResult.created_at.asc()).all()

    # 数据结构：{ date: { domain: count } }
    daily_stats = {}
    all_domains = {} # { domain: total_count }
    
    from urllib.parse import urlparse

    def get_full_domain(url):
        if not url or not url.startswith('http'): return None
        try:
            domain = urlparse(url).netloc.lower()
            if ':' in domain: domain = domain.split(':')[0]
            if domain.startswith('www.'): domain = domain[4:]
            return domain
        except: return None

    def get_top_domain(url):
        domain = get_full_domain(url)
        if not domain: return None
        try:
            parts = domain.split('.')
            if len(parts) <= 2: return domain
            special_suffixes = ['com.cn', 'net.cn', 'gov.cn', 'org.cn', 'edu.cn', 'bj.cn', 'sh.cn']
            last_two = '.'.join(parts[-2:])
            if last_two in special_suffixes and len(parts) >= 3:
                return '.'.join(parts[-3:])
            return '.'.join(parts[-2:])
        except: return domain

    for r in results:
        date_str = r.created_at.strftime('%Y-%m-%d')
        if date_str not in daily_stats:
            daily_stats[date_str] = {}
            
        refs = json.loads(r.references) if r.references else []
        for ref in refs:
            url = ref.get('url', '')
            if level == 'full':
                domain = get_full_domain(url)
            else:
                domain = get_top_domain(url)
                
            if not domain: continue
            
            # 如果指定了域名，则只统计指定的域名
            if selected_domains and domain not in selected_domains:
                continue

            daily_stats[date_str][domain] = daily_stats[date_str].get(domain, 0) + 1
            all_domains[domain] = all_domains.get(domain, 0) + 1

    # 确定要展示的域名列表
    if selected_domains:
        # 使用用户选择的域名（保持顺序或按总数排序）
        display_domains = [d for d in selected_domains if d in all_domains]
        # 如果选择的域名在数据中没出现，也展示出来（数据为0）
        for d in selected_domains:
            if d not in display_domains:
                display_domains.append(d)
    else:
        # 自动获取 Top N 域名
        top_domains = sorted(all_domains.items(), key=lambda x: x[1], reverse=True)[:top_n]
        display_domains = [d[0] for d in top_domains]

    # 构建时间轴
    dates = sorted(daily_stats.keys())
    
    # 构建 Series 数据
    series_data = []
    for domain in display_domains:
        data = []
        for d in dates:
            data.append(daily_stats[d].get(domain, 0))
        series_data.append({
            'name': domain,
            'type': 'line',
            'smooth': True,
            'data': data
        })

    return jsonify({
        'success': True,
        'dates': dates,
        'series': series_data,
        'top_domains': display_domains
    })


# ==================== 商业看板 API ====================

def _safe_json_list(value):
    if not value:
        return []
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, list) else []
    except Exception:
        return []


def _domain_from_url(url):
    if not url:
        return ''
    try:
        from urllib.parse import urlparse
        parsed = urlparse(url if '://' in url else f'https://{url}')
        return (parsed.netloc or parsed.path).lower().replace('www.', '').split('/')[0]
    except Exception:
        return ''


def _build_geo_overview(user_id):
    tasks = MonitorTask.query.filter_by(user_id=user_id).order_by(MonitorTask.created_at.desc()).all()
    task_ids = [task.id for task in tasks]
    results = CollectionResult.query.filter(CollectionResult.task_id.in_(task_ids)).all() if task_ids else []
    manuscripts = GeoManuscript.query.filter_by(user_id=user_id).all()
    config = SentimentConfig.query.filter_by(user_id=user_id, is_default=True).first()

    total_results = len(results)
    exposed_results = sum(1 for result in results if result.has_brand_exposure)
    exposure_rate = round((exposed_results / total_results) * 100, 1) if total_results else 0

    platform_counts = {}
    reference_counts = {}
    question_counts = {}
    sentiment_counts = {'positive': 0, 'neutral': 0, 'negative': 0}
    for result in results:
        platform_counts[result.platform] = platform_counts.get(result.platform, 0) + 1
        question_counts[result.question] = question_counts.get(result.question, 0) + 1
        for ref in _safe_json_list(result.references):
            domain = _domain_from_url(ref.get('url') or ref.get('link') or ref.get('domain') or '')
            if domain:
                reference_counts[domain] = reference_counts.get(domain, 0) + 1
        sentiment = None
        try:
            sentiment = (json.loads(result.ai_sentiment_result or '{}') or {}).get('sentiment')
        except Exception:
            sentiment = None
        if sentiment in sentiment_counts:
            sentiment_counts[sentiment] += 1
        elif result.has_brand_exposure:
            sentiment_counts['positive'] += 1
        else:
            sentiment_counts['neutral'] += 1

    running_tasks = sum(1 for task in tasks if task.status in ['running', 'paused'])
    completed_tasks = sum(1 for task in tasks if task.status == 'completed')
    failed_tasks = sum(1 for task in tasks if task.status == 'failed')
    active_platforms = len(platform_counts)

    top_references = [
        {'domain': domain, 'count': count}
        for domain, count in sorted(reference_counts.items(), key=lambda item: item[1], reverse=True)[:8]
    ]
    platform_mix = [
        {'platform': platform, 'count': count}
        for platform, count in sorted(platform_counts.items(), key=lambda item: item[1], reverse=True)
    ]

    recommendations = []
    if not tasks:
        recommendations.append({
            'level': 'high',
            'title': '创建第一个品牌监测任务',
            'detail': '先选择 3-5 个核心问题和 2-3 个 AI 平台，建立可持续追踪的基线。'
        })
    if tasks and exposure_rate < 40:
        recommendations.append({
            'level': 'high',
            'title': '品牌曝光偏低，优先补齐内容资产',
            'detail': '围绕高频问题发布可引用的解释页、对比页和 FAQ，让 AI 回答更容易提到品牌。'
        })
    if tasks and active_platforms < 3:
        recommendations.append({
            'level': 'medium',
            'title': '扩大平台覆盖',
            'detail': '至少覆盖豆包、DeepSeek、Kimi、千问/元宝中的 3 个平台，避免单平台结论偏差。'
        })
    if total_results and not top_references:
        recommendations.append({
            'level': 'medium',
            'title': '引用来源不足',
            'detail': '当前回答较少返回引用链接，建议加入更容易触发引用的问题类型。'
        })
    if failed_tasks:
        recommendations.append({
            'level': 'medium',
            'title': '处理失败任务',
            'detail': '检查浏览器登录状态、采集间隔和平台限制，避免后续数据断层。'
        })
    if not recommendations:
        recommendations.append({
            'level': 'watch',
            'title': '持续观察趋势变化',
            'detail': '保持每周固定采集，并重点观察品牌曝光率、引用域名和竞品出现频率。'
        })

    ai_ready = bool(config and config.enable_ai_sentiment and config.ai_api_url and config.ai_api_key and config.ai_model_name)
    return {
        'summary': {
            'tasks': len(tasks),
            'running_tasks': running_tasks,
            'completed_tasks': completed_tasks,
            'results': total_results,
            'exposure_rate': exposure_rate,
            'active_platforms': active_platforms,
            'geo_manuscripts': len(manuscripts),
            'reference_domains': len(reference_counts),
        },
        'platform_mix': platform_mix,
        'top_references': top_references,
        'sentiment': sentiment_counts,
        'top_questions': [
            {'question': question, 'count': count}
            for question, count in sorted(question_counts.items(), key=lambda item: item[1], reverse=True)[:6]
        ],
        'recommendations': recommendations,
        'ai_config': {
            'ready': ai_ready,
            'config_id': config.id if config else None,
            'name': config.name if config else '',
            'platform': _ai_api_mode(config) if config else 'openai',
            'api_url': config.ai_api_url if config and config.ai_api_url else '',
            'model': config.ai_model_name if config and config.ai_model_name else '',
        }
    }


@app.route('/api/insights/overview', methods=['GET'])
@login_required
def get_insights_overview():
    """获取商业看板汇总、默认建议和 AI 配置状态。"""
    return jsonify({
        'success': True,
        'data': _build_geo_overview(current_user.id)
    })


def _score_level(score):
    if score >= 80:
        return {'label': '健康', 'type': 'success'}
    if score >= 60:
        return {'label': '可优化', 'type': 'warning'}
    return {'label': '需要关注', 'type': 'danger'}


def _build_insight_scorecard(user_id):
    tasks = MonitorTask.query.filter_by(user_id=user_id).order_by(MonitorTask.created_at.asc()).all()
    task_ids = [task.id for task in tasks]
    results = CollectionResult.query.filter(CollectionResult.task_id.in_(task_ids)).order_by(CollectionResult.created_at.asc()).all() if task_ids else []
    manuscripts = GeoManuscript.query.filter_by(user_id=user_id).all()

    total = len(results)
    exposed = sum(1 for result in results if result.has_brand_exposure)
    exposure_rate = round(exposed * 100 / total, 1) if total else 0

    platform_stats = {}
    daily_stats = {}
    reference_counts = {}
    question_stats = {}
    screenshot_count = 0
    reference_result_count = 0
    competitor_mentions = 0

    task_competitors = {}
    for task in tasks:
        task_competitors[task.id] = [item.lower() for item in _safe_json_list(task.competitor_brands)]

    for result in results:
        date_key = result.created_at.strftime('%Y-%m-%d')
        daily = daily_stats.setdefault(date_key, {'answers': 0, 'exposed': 0})
        daily['answers'] += 1
        if result.has_brand_exposure:
            daily['exposed'] += 1

        platform = platform_stats.setdefault(result.platform, {'platform': result.platform, 'answers': 0, 'exposed': 0})
        platform['answers'] += 1
        if result.has_brand_exposure:
            platform['exposed'] += 1

        question = question_stats.setdefault(result.question, {'question': result.question, 'answers': 0, 'exposed': 0})
        question['answers'] += 1
        if result.has_brand_exposure:
            question['exposed'] += 1

        if result.screenshot_path and _resolve_screenshot_path(result.screenshot_path):
            screenshot_count += 1

        refs = _safe_json_list(result.references)
        if refs:
            reference_result_count += 1
        for ref in refs:
            domain = _domain_from_url(ref.get('url') or ref.get('link') or ref.get('domain') or '')
            if domain:
                reference_counts[domain] = reference_counts.get(domain, 0) + 1

        answer_text = (result.answer or '').lower()
        if any(comp and comp in answer_text for comp in task_competitors.get(result.task_id, [])):
            competitor_mentions += 1

    active_platforms = len(platform_stats)
    reference_domains = len(reference_counts)
    screenshot_rate = round(screenshot_count * 100 / total, 1) if total else 0
    reference_rate = round(reference_result_count * 100 / total, 1) if total else 0
    data_quality = round((screenshot_rate * 0.45) + (reference_rate * 0.35) + (min(total, 20) / 20 * 100 * 0.2), 1) if total else 0
    coverage_score = round(min(active_platforms / 4 * 100, 100), 1)
    source_score = round(min(reference_domains / 8 * 100, 100) * 0.65 + min(len(manuscripts) / 5 * 100, 100) * 0.35, 1)
    geo_score = round(exposure_rate * 0.38 + coverage_score * 0.22 + source_score * 0.2 + data_quality * 0.2, 1)

    trend = []
    for date_key in sorted(daily_stats.keys())[-14:]:
        item = daily_stats[date_key]
        trend.append({
            'date': date_key[5:],
            'answers': item['answers'],
            'exposure_rate': round(item['exposed'] * 100 / item['answers'], 1) if item['answers'] else 0
        })

    platforms = []
    for item in sorted(platform_stats.values(), key=lambda row: row['answers'], reverse=True):
        platforms.append({
            'platform': item['platform'],
            'answers': item['answers'],
            'exposure_rate': round(item['exposed'] * 100 / item['answers'], 1) if item['answers'] else 0
        })

    sources = [
        {'domain': domain, 'count': count}
        for domain, count in sorted(reference_counts.items(), key=lambda entry: entry[1], reverse=True)[:6]
    ]

    weak_questions = []
    for item in sorted(question_stats.values(), key=lambda row: (row['exposed'] / row['answers'] if row['answers'] else 0, -row['answers']))[:5]:
        weak_questions.append({
            'question': item['question'],
            'answers': item['answers'],
            'exposure_rate': round(item['exposed'] * 100 / item['answers'], 1) if item['answers'] else 0
        })

    if not tasks:
        action = {'title': '先创建第一个监测任务', 'detail': '选择 3-5 个真实客户会问的问题，先采集一轮基线数据。'}
    elif not total:
        action = {'title': '先跑一次采集', 'detail': '没有采集结果时，所有分析都只是空壳。先完成一轮采集。'}
    elif exposure_rate < 30:
        action = {'title': '优先解决品牌不出现的问题', 'detail': '围绕曝光率最低的问题，补 FAQ、对比页、案例页，让 AI 有内容可引用。'}
    elif active_platforms < 3:
        action = {'title': '扩大 AI 平台覆盖', 'detail': '至少覆盖 3 个平台，否则结论容易被单个平台偏好带偏。'}
    elif reference_domains < 3:
        action = {'title': '补充可被引用的内容源', 'detail': '让官网、媒体稿、知识页成为 AI 回答里的稳定来源。'}
    else:
        action = {'title': '保持监测节奏', 'detail': '每周固定采集，观察曝光率、引用来源和竞品压力是否变化。'}

    return {
        'score': {'value': geo_score, **_score_level(geo_score)},
        'cards': [
            {'key': 'exposure', 'name': '品牌出现率', 'value': exposure_rate, 'unit': '%', 'hint': 'AI 回答里提到你的比例', 'status': _score_level(exposure_rate)},
            {'key': 'coverage', 'name': '平台覆盖', 'value': active_platforms, 'unit': '个平台', 'hint': '建议至少覆盖 3 个平台', 'status': _score_level(coverage_score)},
            {'key': 'sources', 'name': '引用来源', 'value': reference_domains, 'unit': '个域名', 'hint': 'AI 引用了哪些来源', 'status': _score_level(source_score)},
            {'key': 'quality', 'name': '数据质量', 'value': data_quality, 'unit': '分', 'hint': '截图、引用和样本是否完整', 'status': _score_level(data_quality)},
        ],
        'action': action,
        'charts': {
            'trend': trend,
            'platforms': platforms,
            'sources': sources,
            'weak_questions': weak_questions,
        },
        'explainer': {
            'sample_size': total,
            'screenshot_rate': screenshot_rate,
            'reference_rate': reference_rate,
            'competitor_mentions': competitor_mentions,
        }
    }


@app.route('/api/insights/scorecard', methods=['GET'])
@login_required
def get_insights_scorecard():
    """获取小白友好的核心分析看板数据。"""
    return jsonify({'success': True, 'data': _build_insight_scorecard(current_user.id)})


@app.route('/api/insights/ai-analysis', methods=['POST'])
@login_required
def run_ai_insights_analysis():
    """基于采集数据调用用户配置的 AI API，生成观察和下一步动作。"""
    overview = _build_geo_overview(current_user.id)
    config = SentimentConfig.query.filter_by(user_id=current_user.id, is_default=True).first()
    if not (config and config.enable_ai_sentiment and config.ai_api_url and config.ai_api_key and config.ai_model_name):
        return jsonify({
            'success': False,
            'needs_config': True,
            'message': '请先在 AI 分析设置里填写 API URL、API Key 和模型名称，并设为默认配置。',
            'fallback': overview['recommendations']
        }), 400

    prompt = f"""你是 GEO 品牌监测分析师。请基于以下 AI 平台采集结果汇总，输出 JSON。

数据汇总：
{json.dumps(overview, ensure_ascii=False)}

请严格输出：
{{
  "summary": "一句话总览",
  "observations": ["3条关键观察"],
  "actions": ["3条下一步动作，具体可执行"],
  "risks": ["1-3条风险提醒"],
  "experiments": ["1-3条下一轮测试问题或内容实验"]
}}
"""
    try:
        import requests
        mode = _ai_api_mode(config)
        if mode == 'anthropic':
            api_url = _normalize_anthropic_messages_url(config.ai_api_url)
            headers = {
                'Content-Type': 'application/json',
                'x-api-key': config.ai_api_key,
                'anthropic-version': '2023-06-01',
            }
            payload = {
                'model': config.ai_model_name,
                'max_tokens': 1200,
                'temperature': 0.2,
                'messages': [{'role': 'user', 'content': prompt}],
            }
        else:
            api_url = _normalize_openai_chat_url(config.ai_api_url)
            headers = {
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {config.ai_api_key}'
            }
            payload = {
                'model': config.ai_model_name,
                'messages': [{'role': 'user', 'content': prompt}],
                'temperature': 0.2,
                'max_tokens': 1200
            }
        response = requests.post(api_url, headers=headers, json=payload, timeout=75)
        response.raise_for_status()
        data = response.json()
        if mode == 'anthropic':
            raw = ''.join(
                block.get('text', '')
                for block in data.get('content', [])
                if isinstance(block, dict) and block.get('type') == 'text'
            )
        else:
            raw = data.get('choices', [{}])[0].get('message', {}).get('content', '')
        parsed = json.loads(_extract_json_text(raw), strict=False)
        return jsonify({'success': True, 'analysis': parsed, 'overview': overview, 'api_mode': mode})
    except requests.exceptions.HTTPError as e:
        status_code = e.response.status_code if e.response is not None else 'unknown'
        body = ''
        if e.response is not None:
            try:
                body = e.response.text[:500]
            except Exception:
                body = ''
        logger.warning(f"AI看板分析失败: HTTP {status_code} {body}")
        return jsonify({
            'success': False,
            'message': f'AI 分析失败：接口返回 HTTP {status_code}。请检查 API URL、Key、模型名称和账户余额。',
            'detail': body,
            'api_mode': _ai_api_mode(config),
            'api_url': _normalize_anthropic_messages_url(config.ai_api_url) if _ai_api_mode(config) == 'anthropic' else _normalize_openai_chat_url(config.ai_api_url),
            'fallback': overview['recommendations']
        }), 502
    except Exception as e:
        logger.warning(f"AI看板分析失败: {e}")
        return jsonify({
            'success': False,
            'message': f'AI 分析失败：{e}',
            'api_mode': _ai_api_mode(config),
            'api_url': _normalize_anthropic_messages_url(config.ai_api_url) if _ai_api_mode(config) == 'anthropic' else _normalize_openai_chat_url(config.ai_api_url),
            'fallback': overview['recommendations']
        }), 502


# ==================== 舆情设置 API ====================

@app.route('/api/sentiment/configs', methods=['GET'])
@login_required
def get_sentiment_configs():
    """获取用户的舆情配置列表"""
    configs = SentimentConfig.query.filter_by(user_id=current_user.id).all()
    return jsonify({
        'success': True,
        'configs': [c.to_dict() for c in configs]
    })


@app.route('/api/sentiment/configs/<int:config_id>', methods=['GET'])
@login_required
def get_sentiment_config(config_id):
    """获取单个舆情配置"""
    config = SentimentConfig.query.filter_by(id=config_id, user_id=current_user.id).first()
    if not config:
        return jsonify({'success': False, 'message': '配置不存在'}), 404
    return jsonify({'success': True, 'config': config.to_dict()})


@app.route('/api/sentiment/configs', methods=['POST'])
@login_required
def create_sentiment_config():
    """创建舆情配置"""
    data = request.get_json()
    
    # 如果设置为默认，取消其他默认配置
    if data.get('is_default'):
        SentimentConfig.query.filter_by(user_id=current_user.id, is_default=True).update({'is_default': False})
    
    config = SentimentConfig(
        user_id=current_user.id,
        name=data.get('name', ''),
        positive_words=json.dumps(data.get('positive_words', [])),
        negative_words=json.dumps(data.get('negative_words', [])),
        enable_ai_sentiment=data.get('enable_ai_sentiment', False),
        ai_platform=data.get('ai_platform'),
        ai_api_url=data.get('ai_api_url'),
        ai_api_key=data.get('ai_api_key'),
        ai_model_name=data.get('ai_model_name'),
        ai_prompt=data.get('ai_prompt'),
        is_default=data.get('is_default', False)
    )
    
    db.session.add(config)
    db.session.commit()
    
    return jsonify({'success': True, 'config': config.to_dict()})


@app.route('/api/sentiment/configs/<int:config_id>', methods=['PUT'])
@login_required
def update_sentiment_config(config_id):
    """更新舆情配置"""
    config = SentimentConfig.query.filter_by(id=config_id, user_id=current_user.id).first()
    if not config:
        return jsonify({'success': False, 'message': '配置不存在'}), 404
    
    data = request.get_json()
    
    # 如果设置为默认，取消其他默认配置
    if data.get('is_default') and not config.is_default:
        SentimentConfig.query.filter_by(user_id=current_user.id, is_default=True).update({'is_default': False})
    
    config.name = data.get('name', config.name)
    config.positive_words = json.dumps(data.get('positive_words', [])) if 'positive_words' in data else config.positive_words
    config.negative_words = json.dumps(data.get('negative_words', [])) if 'negative_words' in data else config.negative_words
    config.enable_ai_sentiment = data.get('enable_ai_sentiment', config.enable_ai_sentiment)
    config.ai_platform = data.get('ai_platform', config.ai_platform)
    config.ai_api_url = data.get('ai_api_url', config.ai_api_url)
    config.ai_api_key = data.get('ai_api_key', config.ai_api_key)
    config.ai_model_name = data.get('ai_model_name', config.ai_model_name)
    config.ai_prompt = data.get('ai_prompt', config.ai_prompt)
    config.is_default = data.get('is_default', config.is_default)
    
    db.session.commit()
    
    return jsonify({'success': True, 'config': config.to_dict()})


@app.route('/api/sentiment/configs/<int:config_id>', methods=['DELETE'])
@login_required
def delete_sentiment_config(config_id):
    """删除舆情配置"""
    config = SentimentConfig.query.filter_by(id=config_id, user_id=current_user.id).first()
    if not config:
        return jsonify({'success': False, 'message': '配置不存在'}), 404
    
    db.session.delete(config)
    db.session.commit()
    
    return jsonify({'success': True, 'message': '删除成功'})


@app.route('/api/sentiment/analyze', methods=['POST'])
@login_required
def analyze_sentiment():
    """分析舆情（普通舆情+智能舆情）"""
    data = request.get_json()
    text = data.get('text', '')
    keyword = data.get('keyword', '')
    config_id = data.get('config_id')
    
    # 获取配置
    if config_id:
        config = SentimentConfig.query.filter_by(id=config_id, user_id=current_user.id).first()
    else:
        # 获取默认配置
        config = SentimentConfig.query.filter_by(user_id=current_user.id, is_default=True).first()
    
    if not config:
        # 如果没有配置，使用默认关键词
        positive_words = [
            '好', '不错', '优秀', '棒', '赞', '推荐', '喜欢', '满意', '靠谱',
            '好评', '出色', '卓越', '完美', '理想', '认可', '信赖', '放心',
            '值得', '划算', '实惠', '优质', '高端', '专业', '有效', '神奇', '见效',
            '吸收', '补水', '保湿', '美白', '抗衰', '修复', '温和', '清爽', '不油腻',
            '无刺激', '抗过敏', '口碑好', '效果好', '质量好', '性价比高', '回购',
            '正品', '安全', '健康', '自然', '纯正', '地道', '正宗'
        ]
        negative_words = [
            '差', '不好', '烂', '垃圾', '失望', '后悔', '差评', '糟糕', '骗人',
            '虚假', '没效果', '无用', '无效', '过敏', '刺激', '油腻', '干燥',
            '紧绷', '爆痘', '泛红', '瘙痒', '副作用', '假货', '劣质', '粗糙',
            '廉价', '智商税', '坑人', '受骗', '上当', '不推荐', '避雷',
            '踩坑', '鸡肋', '不值', '浪费', '昂贵', '伤皮肤', '有问题',
            '不合格', '超标', '有害', '致癌', '激素', '荧光剂', '重金属'
        ]
        enable_ai_sentiment = False
        ai_prompt = ''
    else:
        positive_words = json.loads(config.positive_words) if config.positive_words else []
        negative_words = json.loads(config.negative_words) if config.negative_words else []
        enable_ai_sentiment = config.enable_ai_sentiment
        ai_prompt = config.ai_prompt
    
    # 普通舆情分析
    local_result = analyze_sentiment_local(text, keyword, positive_words, negative_words)
    
    # 智能舆情分析
    ai_result = None
    if enable_ai_sentiment and config and config.ai_api_url and config.ai_api_key:
        ai_result = analyze_sentiment_ai(text, keyword, config)
    
    return jsonify({
        'success': True,
        'local_sentiment': local_result,
        'ai_sentiment': ai_result,
        'config_id': config.id if config else None
    })


def analyze_sentiment_local(text, keyword, positive_words, negative_words):
    """本地舆情分析（基于关键词匹配）"""
    if not text:
        return {'sentiment': 'neutral', 'score': 0, 'label': '中性', 'color': '#8c8c8c', 'matched_words': []}
    
    # 如果指定了关键词，只分析关键词附近的文本
    if keyword:
        kw_index = text.find(keyword)
        if kw_index != -1:
            start = max(0, kw_index - 50)
            end = min(len(text), kw_index + len(keyword) + 50)
            text_segment = text[start:end]
        else:
            text_segment = text[:100]
    else:
        text_segment = text[:200]
    
    # 检查是否包含价格信息（如"10万"、"15.5万"等），如果有则判定为负面
    import re
    price_pattern = r'\d+(?:\.\d+)?\s*万'
    if re.search(price_pattern, text_segment):
        return {'sentiment': 'negative', 'score': -0.8, 'label': '负面', 'color': '#f5222d', 'matched_words': ['价格信息']}
    
    positive_score = 0
    negative_score = 0
    matched_positive = []
    matched_negative = []
    
    for word in positive_words:
        count = text_segment.count(word)
        if count > 0:
            positive_score += count
            matched_positive.append(word)
    
    for word in negative_words:
        count = text_segment.count(word)
        if count > 0:
            negative_score += count
            matched_negative.append(word)
    
    total = positive_score + negative_score
    if total == 0:
        return {'sentiment': 'neutral', 'score': 0, 'label': '中性', 'color': '#8c8c8c', 'matched_words': []}
    
    ratio = (positive_score - negative_score) / total
    
    if ratio > 0.3:
        return {
            'sentiment': 'positive',
            'score': ratio,
            'label': '正面',
            'color': '#52c41a',
            'matched_words': {'positive': matched_positive, 'negative': matched_negative}
        }
    elif ratio < -0.3:
        return {
            'sentiment': 'negative',
            'score': ratio,
            'label': '负面',
            'color': '#ff4d4f',
            'matched_words': {'positive': matched_positive, 'negative': matched_negative}
        }
    else:
        return {
            'sentiment': 'neutral',
            'score': ratio,
            'label': '中性',
            'color': '#fa8c16',
            'matched_words': {'positive': matched_positive, 'negative': matched_negative}
        }


def analyze_sentiment_ai(text, keyword, config):
    """智能舆情分析（调用第三方API）"""
    try:
        import requests
        
        # 构建prompt
        if not config.ai_prompt:
            prompt = f"""请分析以下文本中关于"{keyword if keyword else '内容'}"的舆情倾向：
            
文本：{text[:500]}

请判断是正面、负面还是中性，并给出简短理由。
输出格式：{{"sentiment": "positive|negative|neutral", "score": -1到1之间的数值, "label": "正面|负面|中性", "reason": "分析理由"}}
"""
        else:
            # 转义输出格式中的大括号，避免format解析错误
            escaped_prompt = config.ai_prompt.replace('{', '{{').replace('}', '}}')
            # 然后恢复text和keyword的占位符
            escaped_prompt = escaped_prompt.replace('{{text}}', '{text}').replace('{{keyword}}', '{keyword}')
            prompt = escaped_prompt.format(text=text[:1000], keyword=keyword if keyword else '')
        
        # 调用API
        headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {config.ai_api_key}'
        }
        
        payload = {
            'model': config.ai_model_name,
            'messages': [{'role': 'user', 'content': prompt}],
            'temperature': 0.3
        }
        
        logger.info(f"调用AI舆情分析API: {config.ai_api_url}")
        logger.info(f"使用模型: {config.ai_model_name}")
        
        response = requests.post(config.ai_api_url, headers=headers, json=payload, timeout=30)
        response.raise_for_status()
        
        result = response.json()
        ai_response = result.get('choices', [{}])[0].get('message', {}).get('content', '')
        
        logger.info(f"AI返回内容: {ai_response[:300]}")
        
        # 解析AI返回结果
        try:
            # 尝试清理AI返回的内容（去除可能的markdown代码块标记）
            cleaned_response = ai_response.strip()
            if cleaned_response.startswith('```json'):
                cleaned_response = cleaned_response[7:]
            if cleaned_response.endswith('```'):
                cleaned_response = cleaned_response[:-3]
            cleaned_response = cleaned_response.strip()
            
            parsed = json.loads(cleaned_response)
            return {
                'sentiment': parsed.get('sentiment', 'neutral'),
                'score': parsed.get('score', 0),
                'label': parsed.get('label', '中性'),
                'reason': parsed.get('reason', ''),
                'raw_response': ai_response
            }
        except json.JSONDecodeError as e:
            # JSON解析失败，记录详细错误信息
            logger.error(f"AI舆情分析JSON解析失败: {str(e)}")
            logger.error(f"原始响应: {ai_response[:500]}")
            return {
                'sentiment': 'neutral',
                'score': 0,
                'label': '中性',
                'reason': f'分析失败: JSON解析错误',
                'raw_response': ai_response
            }
        except Exception as e:
            # 其他解析失败
            logger.error(f"AI舆情分析解析失败: {str(e)}")
            return {
                'sentiment': 'neutral',
                'score': 0,
                'label': '中性',
                'reason': f'分析失败: {str(e)}',
                'raw_response': ai_response
            }
    
    except Exception as e:
        logger.error(f"智能舆情分析失败: {str(e)}")
        logger.error(f"API URL: {config.ai_api_url}")
        logger.error(f"模型: {config.ai_model_name}")
        logger.error("API Key状态: %s", "已配置" if config.ai_api_key else "未配置")
        
        # 提供更友好的错误提示
        error_msg = str(e)
        if '403' in error_msg:
            if 'balance' in error_msg.lower() or 'insufficient' in error_msg.lower():
                error_msg = '账户余额不足，请充值后重试，或切换到更便宜的模型（如 Qwen/Qwen2-7B-Instruct）'
            else:
                error_msg = 'API鉴权失败，请检查API密钥是否正确、余额是否充足、模型是否有访问权限'
        elif '404' in error_msg:
            error_msg = 'API地址不存在，请检查API地址是否正确'
        elif 'timeout' in error_msg.lower():
            error_msg = 'API请求超时，请检查网络连接或稍后重试'
        
        return {
            'sentiment': 'neutral',
            'score': 0,
            'label': '中性',
            'reason': f'分析失败: {error_msg}',
            'error': str(e)
        }


# ==================== 初始化数据库 ====================

@app.cli.command()
def init_db():
    """初始化数据库"""
    db.create_all()
    print("数据库初始化完成！")


@app.cli.command()
def create_admin():
    """创建管理员账户"""
    admin = User(username='admin', email='admin@example.com')
    admin.set_password('admin')
    db.session.add(admin)
    db.session.commit()
    print("管理员账户创建成功！")
    print("用户名: admin")
    print("密码: admin")


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.debug = False
    app.use_debugger = False
    app.use_reloader = False
    app.run(host='0.0.0.0', port=6001)
