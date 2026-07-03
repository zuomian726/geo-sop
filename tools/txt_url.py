import openpyxl
import re
from pathlib import Path

# Load workbook
file_path = r'c:\Users\houch\Desktop\pythonTools\Excel\aiva\监控-赛豆词包_GEO效果_2026-05-30_2026-06-05_11.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)

# AI回答 sheets to process
ai_sheets = [name for name in wb.sheetnames if 'AI回答' in name]
print(f"Processing sheets: {ai_sheets}")

for sheet_name in ai_sheets:
    ws = wb[sheet_name]
    print(f"\n=== Processing {sheet_name} ===")

    # Find columns that match date pattern YYYY-MM-DD
    date_cols = []
    date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
    for cell in ws[1]:
        if cell.value and date_pattern.match(str(cell.value)):
            date_cols.append(cell.column)
            print(f"Found date column: {cell.column_letter} ({cell.value})")

    print(f"Total date columns found: {len(date_cols)}")

    # Process each row
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in date_cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                text = str(cell.value)
                # Extract URLs from text
                urls = re.findall(r'https?://[^\s<>"\')\]]+', text)
                if urls:
                    # Replace with URLs separated by newlines
                    print(f"Row {row_idx}, Col {col_idx}: Found URLs: {urls}")
                    cell.value = '\n'.join(urls)
                else:
                    # Replace with empty string if no URLs found
                    print(f"Row {row_idx}, Col {col_idx}: No URLs found, setting to empty")
                    cell.value = ''

# Save the modified workbook
output_path = file_path
wb.save(output_path)
print(f"\nSaved to: {output_path}")